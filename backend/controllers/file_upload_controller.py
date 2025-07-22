# controllers/file_upload_controller.py - COMPLETE ENHANCED VERSION

import os
import time
import uuid
from pathlib import Path
from flask import Blueprint, request, jsonify, send_file, Response, send_from_directory
from flask_jwt_extended import jwt_required, get_jwt_identity
from werkzeug.utils import secure_filename
from werkzeug.datastructures import FileStorage
from typing import List, Dict, Any, Tuple
from models.user import User
from models.file_analysis import FileAnalysis, FileAnalysisCreate
from services.material_analysis import MaterialAnalysisService, CostEstimationService
from services.step_renderer import StepRendererEnhanced
from services.cad_converter import cad_converter, get_file_type_enhanced, needs_step_conversion  # ✅ YENİ
import numpy as np
import math
import threading
import queue
import re
from difflib import SequenceMatcher

# Blueprint oluştur
upload_bp = Blueprint('upload', __name__, url_prefix='/api/upload')

# Konfigürasyon
UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'step', 'stp', 'prt', 'catpart'}  # ✅ YENİ: PRT ve CATPART eklendi
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
MAX_FILES_PER_REQUEST = 100

# Upload klasörünü oluştur
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs("static", exist_ok=True)


class OptimizedBackgroundProcessor:
    def __init__(self):
        self.task_queue = queue.Queue()
        self.results = {}
        self.worker_thread = threading.Thread(target=self._worker, daemon=True)
        self.worker_thread.start()
        
    def _worker(self):
        while True:
            try:
                task = self.task_queue.get(timeout=1)
                if task:
                    task_id, func, args, kwargs = task
                    try:
                        result = func(*args, **kwargs)
                        self.results[task_id] = {"success": True, "result": result}
                    except Exception as e:
                        self.results[task_id] = {"success": False, "error": str(e)}
                    self.task_queue.task_done()
            except queue.Empty:
                continue
                
    def add_task(self, func, args=(), kwargs=None):
        task_id = str(uuid.uuid4())
        self.task_queue.put((task_id, func, args, kwargs or {}))
        return task_id
        
    def get_result(self, task_id):
        return self.results.get(task_id)
    
# Global background processor
bg_processor = OptimizedBackgroundProcessor()

# ===== PDF-STEP MATCHING UTILITIES =====

def normalize_filename(filename: str) -> str:
    """Dosya adını normalize et (karşılaştırma için)"""
    # Dosya uzantısını kaldır
    name = os.path.splitext(filename)[0]
    # Küçük harfe çevir, özel karakterleri kaldır
    normalized = re.sub(r'[^a-zA-Z0-9]', '', name.lower())
    return normalized

def extract_numbers_from_filename(filename: str) -> List[str]:
    """Dosya adından sayıları çıkar"""
    numbers = re.findall(r'\d+', filename)
    return numbers

def calculate_filename_similarity(name1: str, name2: str) -> float:
    """İki dosya adı arasındaki benzerlik oranını hesapla"""
    # Normalize edilmiş adları karşılaştır
    norm1 = normalize_filename(name1)
    norm2 = normalize_filename(name2)
    
    # Sequence matcher ile genel benzerlik
    similarity = SequenceMatcher(None, norm1, norm2).ratio()
    
    # Sayısal benzerlik kontrolü
    numbers1 = extract_numbers_from_filename(name1)
    numbers2 = extract_numbers_from_filename(name2)
    
    # Ortak sayılar varsa bonus ver
    if numbers1 and numbers2:
        common_numbers = set(numbers1) & set(numbers2)
        if common_numbers:
            # En büyük ortak sayıyı bul
            max_common = max(common_numbers, key=len) if common_numbers else ""
            if len(max_common) >= 3:  # En az 3 haneli sayı
                similarity += 0.3  # Bonus
    
    return min(similarity, 1.0)  # 1.0'ı geçmesin

def match_pdf_to_cad_files(pdf_files: List[Dict], cad_files: List[Dict]) -> List[Dict]:
    """✅ ENHANCED - PDF dosyalarını CAD dosyalarıyla (STEP/PRT/CATPART) eşleştir"""
    matches = []
    
    for pdf_info in pdf_files:
        pdf_filename = pdf_info['original_filename']
        best_match = None
        best_score = 0.0
        
        # Her CAD dosyası ile karşılaştır
        for cad_info in cad_files:
            cad_filename = cad_info['original_filename']
            score = calculate_filename_similarity(pdf_filename, cad_filename)
            
            if score > best_score:
                best_score = score
                best_match = cad_info
        
        # Eşleştirme sonucunu kaydet
        match_result = {
            "pdf_file": pdf_info,
            "cad_file": best_match,  # ✅ STEP yerine CAD
            "match_score": round(best_score * 100, 1),
            "match_quality": get_match_quality(best_score),
            "analysis_strategy": determine_analysis_strategy(pdf_info, best_match, best_score)
        }
        
        matches.append(match_result)
        
        cad_type = "UNKNOWN"
        if best_match:
            cad_type = best_match.get('conversion_info', {}).get('original_format', 'STEP').upper()
        
        print(f"[MATCH] 📄 {pdf_filename} ↔ {best_match['original_filename'] if best_match else 'None'} "
              f"({cad_type}) (Score: {match_result['match_score']}% - {match_result['match_quality']})")
    
    return matches

def match_pdf_to_step_files(pdf_files: List[Dict], step_files: List[Dict]) -> List[Dict]:
    """PDF dosyalarını STEP dosyalarıyla eşleştir"""
    matches = []
    
    for pdf_info in pdf_files:
        pdf_filename = pdf_info['original_filename']
        best_match = None
        best_score = 0.0
        
        # Her STEP dosyası ile karşılaştır
        for step_info in step_files:
            step_filename = step_info['original_filename']
            score = calculate_filename_similarity(pdf_filename, step_filename)
            
            if score > best_score:
                best_score = score
                best_match = step_info
        
        # Eşleştirme sonucunu kaydet
        match_result = {
            "pdf_file": pdf_info,
            "step_file": best_match,
            "match_score": round(best_score * 100, 1),
            "match_quality": get_match_quality(best_score),
            "analysis_strategy": determine_analysis_strategy(pdf_info, best_match, best_score)
        }
        
        matches.append(match_result)
        
        print(f"[MATCH] 📄 {pdf_filename} ↔ {best_match['original_filename'] if best_match else 'None'} "
              f"(Score: {match_result['match_score']}% - {match_result['match_quality']})")
    
    return matches

def get_match_quality(score: float) -> str:
    """Eşleştirme kalitesini belirle"""
    if score >= 0.8:
        return "Excellent"
    elif score >= 0.6:
        return "Good"
    elif score >= 0.4:
        return "Fair"
    elif score >= 0.2:
        return "Poor"
    else:
        return "None"

def determine_analysis_strategy(pdf_info: Dict, cad_info: Dict, score: float) -> str:
    """✅ ENHANCED - Analiz stratejisini belirle (CAD desteği ile)"""
    if cad_info and score >= 0.6:
        return "pdf_with_matched_cad"  # ✅ CAD desteği
    elif cad_info and score >= 0.3:
        return "pdf_with_possible_cad"  # ✅ CAD desteği
    else:
        return "pdf_only_extract_step"

# ===== HELPER FUNCTIONS =====

def get_current_user():
    """Mevcut kullanıcıyı getir"""
    current_user_id = get_jwt_identity()
    return User.find_by_id(current_user_id)

def allowed_file(filename: str) -> bool:
    """✅ ENHANCED - PRT/CATPART desteği ile dosya uzantısı kontrolü"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_file_type(filename: str) -> str:
    """✅ ENHANCED - PRT/CATPART desteği ile dosya tipini belirle"""
    return get_file_type_enhanced(filename)

def save_uploaded_file(file: FileStorage, upload_folder: str) -> Dict[str, Any]:
    """✅ ENHANCED - Yüklenen dosyayı güvenli şekilde kaydet + CAD conversion info"""
    # Dosya boyutu kontrolü
    file.stream.seek(0, 2)
    file_size = file.stream.tell()
    file.stream.seek(0)
    
    if file_size > MAX_FILE_SIZE:
        raise ValueError(f"Dosya çok büyük. Maksimum boyut: {MAX_FILE_SIZE // (1024*1024)}MB")
    
    # Benzersiz dosya adı oluştur
    original_filename = file.filename
    timestamp = int(time.time())
    unique_filename = f"{timestamp}_{uuid.uuid4().hex[:8]}_{secure_filename(original_filename)}"
    
    # Dosyayı kaydet
    file_path = os.path.join(upload_folder, unique_filename)
    file.save(file_path)
    
    file_type = get_file_type(original_filename)
    
    # ✅ YENİ - CAD conversion bilgisi ekle
    conversion_info = {
        "needs_conversion": needs_step_conversion(original_filename),
        "is_cad_file": file_type in ['step', 'cad_part'],
        "original_format": original_filename.rsplit('.', 1)[1].lower() if '.' in original_filename else 'unknown'
    }
    
    return {
        "original_filename": original_filename,
        "saved_filename": unique_filename,
        "file_path": file_path,
        "file_size": file_size,
        "file_type": file_type,
        "conversion_info": conversion_info  # ✅ YENİ
    }

def convert_cad_to_step_if_needed(file_info: Dict[str, Any]) -> Dict[str, Any]:
    """
    ✅ YENİ - CAD dosyasını gerekirse STEP'e çevirir
    
    Args:
        file_info: save_uploaded_file'dan dönen dosya bilgisi
        
    Returns:
        Dict: Çevirme sonucu ve güncellenmiş dosya bilgisi
    """
    try:
        if not file_info.get("conversion_info", {}).get("needs_conversion", False):
            # Çevirmeye gerek yok
            return {
                "success": True,
                "conversion_needed": False,
                "step_file_path": file_info["file_path"],
                "step_file_info": file_info,
                "message": "No conversion needed"
            }
        
        print(f"[CAD-CONVERT] 🔄 Converting {file_info['original_filename']} to STEP...")
        
        # Çevirme işlemi
        original_path = file_info["file_path"]
        
        # STEP dosyası için hedef yol oluştur
        step_filename = file_info["saved_filename"].rsplit('.', 1)[0] + "_converted.step"
        step_file_path = os.path.join(os.path.dirname(original_path), step_filename)
        
        # CAD Converter kullanarak çevir
        conversion_result = cad_converter.convert_to_step(original_path, step_file_path)
        
        if conversion_result["success"]:
            # Çevirme başarılı
            step_file_size = os.path.getsize(step_file_path) if os.path.exists(step_file_path) else 0
            
            # STEP dosyası için yeni file_info oluştur
            step_file_info = {
                "original_filename": file_info["original_filename"],  # Orijinal adı koru
                "saved_filename": step_filename,
                "file_path": step_file_path,
                "file_size": step_file_size,
                "file_type": "step",  # Artık STEP
                "conversion_info": {
                    "needs_conversion": False,
                    "is_cad_file": True,
                    "original_format": file_info["conversion_info"]["original_format"],
                    "converted_from": file_info["conversion_info"]["original_format"],
                    "conversion_time": conversion_result.get("processing_time", 0)
                },
                "source_file_info": file_info  # Orijinal dosya bilgisini sakla
            }
            
            print(f"[CAD-CONVERT] ✅ Conversion successful: {original_path} -> {step_file_path}")
            
            return {
                "success": True,
                "conversion_needed": True,
                "step_file_path": step_file_path,
                "step_file_info": step_file_info,
                "original_file_path": original_path,
                "original_file_info": file_info,
                "processing_time": conversion_result.get("processing_time", 0),
                "message": f"Successfully converted {file_info['conversion_info']['original_format']} to STEP"
            }
        else:
            # Çevirme başarısız
            error_msg = conversion_result.get("error", "Unknown conversion error")
            print(f"[CAD-CONVERT] ❌ Conversion failed: {error_msg}")
            
            return {
                "success": False,
                "conversion_needed": True,
                "error": error_msg,
                "step_file_path": None,
                "step_file_info": None,
                "original_file_path": original_path,
                "original_file_info": file_info,
                "message": f"Conversion failed: {error_msg}"
            }
            
    except Exception as e:
        print(f"[CAD-CONVERT] ❌ Conversion exception: {str(e)}")
        return {
            "success": False,
            "conversion_needed": True,
            "error": f"Conversion exception: {str(e)}",
            "step_file_path": None,
            "step_file_info": None,
            "original_file_path": file_info.get("file_path"),
            "original_file_info": file_info,
            "message": f"Conversion failed with exception: {str(e)}"
        }

# ===== UPLOAD ENDPOINTS =====

@upload_bp.route('/single', methods=['POST'])
@jwt_required()
def upload_single_file():
    """Tek dosya yükleme - OPTIMIZED"""
    try:
        current_user = get_current_user()
        
        if 'file' not in request.files:
            return jsonify({"success": False, "message": "Dosya bulunamadı"}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({"success": False, "message": "Dosya seçilmedi"}), 400
        
        if not allowed_file(file.filename):
            return jsonify({
                "success": False,
                "message": f"Desteklenmeyen dosya türü. İzin verilen: {', '.join(ALLOWED_EXTENSIONS)}"
            }), 400
        
        # Dosyayı kaydet
        file_info = save_uploaded_file(file, UPLOAD_FOLDER)
        
        # Analiz kaydı oluştur
        analysis_record = FileAnalysis.create_analysis({
            "user_id": current_user['id'],
            "filename": file_info['saved_filename'],
            "original_filename": file_info['original_filename'],
            "file_type": file_info['file_type'],
            "file_size": file_info['file_size'],
            "file_path": file_info['file_path'],
            "analysis_status": "uploaded"
        })
        
        return jsonify({
            "success": True,
            "message": "Dosya başarıyla yüklendi",
            "file_info": {
                "analysis_id": analysis_record['id'],
                "filename": file_info['saved_filename'],
                "original_filename": file_info['original_filename'],
                "file_type": file_info['file_type'],
                "file_size": file_info['file_size'],
                "upload_time": analysis_record['created_at']
            }
        }), 201
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Upload error: {str(e)}"
        }), 500

@upload_bp.route('/multiple', methods=['POST'])
@jwt_required()
def upload_multiple_files_with_physical_cad_conversion():
    """✅ ENHANCED - Çoklu dosya yükleme + PDF-CAD eşleştirme + Fiziksel CAD conversion"""
    try:
        current_user = get_current_user()
        
        files = request.files.getlist('files')
        
        if not files or len(files) == 0:
            return jsonify({
                "success": False,
                "message": "Hiç dosya bulunamadı"
            }), 400
        
        if len(files) > MAX_FILES_PER_REQUEST:
            return jsonify({
                "success": False,
                "message": f"Çok fazla dosya. Maksimum: {MAX_FILES_PER_REQUEST}"
            }), 400
        
        print(f"[MULTIPLE-UPLOAD-PHYSICAL] 📁 {len(files)} dosya yükleniyor (Fiziksel CAD conversion aktif)")
        
        # ✅ ENHANCED - DOSYALARI TÜRE GÖRE AYIR (CAD desteği ile)
        pdf_files = []
        cad_files = []  # ✅ STEP + Converted PRT + Converted CATPART
        other_files = []
        failed_uploads = []
        conversion_results = []
        physical_conversion_log = []
        
        for file in files:
            try:
                if file.filename == '':
                    failed_uploads.append({
                        "filename": "unknown",
                        "error": "Boş dosya adı"
                    })
                    continue
                
                if not allowed_file(file.filename):
                    failed_uploads.append({
                        "filename": file.filename,
                        "error": "Desteklenmeyen dosya türü"
                    })
                    continue
                
                # Dosyayı kaydet
                file_info = save_uploaded_file(file, UPLOAD_FOLDER)
                
                # ✅ YENİ - Fiziksel CAD conversion kontrolü
                if file_info.get("conversion_info", {}).get("needs_conversion", False):
                    print(f"[MULTIPLE-UPLOAD-PHYSICAL] 🔄 Fiziksel conversion başlıyor: {file_info['original_filename']}")
                    print(f"[MULTIPLE-UPLOAD-PHYSICAL] 📂 Input file path: {file_info['file_path']}")
                    print(f"[MULTIPLE-UPLOAD-PHYSICAL] 📋 File exists: {os.path.exists(file_info['file_path'])}")
                    
                    # Dosya varlık kontrolü
                    if not os.path.exists(file_info["file_path"]):
                        failed_uploads.append({
                            "filename": file_info['original_filename'],
                            "error": f"Input file not found: {file_info['file_path']}"
                        })
                        print(f"[MULTIPLE-UPLOAD-PHYSICAL] ❌ Input file not found: {file_info['file_path']}")
                        continue
                    
                    # Fiziksel CAD converter kullan
                    try:
                        physical_conversion_result = cad_converter.convert_to_step_with_save(
                            file_info["file_path"],
                            custom_output_name=f"converted_{Path(file_info['original_filename']).stem}",
                            save_original=True
                        )
                    except Exception as conv_error:
                        physical_conversion_result = {
                            "success": False,
                            "error": f"Conversion exception: {str(conv_error)}",
                            "input_path": file_info["file_path"]
                        }
                        print(f"[MULTIPLE-UPLOAD-PHYSICAL] ❌ Conversion exception: {conv_error}")
                        import traceback
                        print(f"[MULTIPLE-UPLOAD-PHYSICAL] 📋 Traceback: {traceback.format_exc()}")
                    
                    conversion_results.append(physical_conversion_result)
                    
                    # Debug logging for conversion result
                    print(f"[MULTIPLE-UPLOAD-PHYSICAL] 📊 Conversion result keys: {list(physical_conversion_result.keys())}")
                    if not physical_conversion_result.get("success", False):
                        print(f"[MULTIPLE-UPLOAD-PHYSICAL] ❌ Conversion error details: {physical_conversion_result.get('error', 'Unknown')}")
                        if "traceback" in physical_conversion_result:
                            print(f"[MULTIPLE-UPLOAD-PHYSICAL] 📋 Conversion traceback: {physical_conversion_result['traceback'][:500]}...")
                    
                    # Fiziksel conversion log'u
                    physical_conversion_log.append({
                        "original_file": file_info['original_filename'],
                        "original_format": file_info.get('conversion_info', {}).get('original_format', 'unknown'),
                        "detected_format": physical_conversion_result.get('detected_format', 'Unknown'),
                        "format_confidence": physical_conversion_result.get('format_confidence', 0),
                        "conversion_successful": physical_conversion_result.get("success", False),
                        "output_path": physical_conversion_result.get("output_path"),
                        "original_saved_path": physical_conversion_result.get("original_saved_path"),
                        "log_file": physical_conversion_result.get("log_file"),
                        "processing_time": physical_conversion_result.get("processing_time", 0),
                        "message": physical_conversion_result.get("message", ""),
                        "error": physical_conversion_result.get("error") if not physical_conversion_result.get("success") else None,
                        "recommendations": physical_conversion_result.get("recommendations", [])
                    })
                    
                    if physical_conversion_result.get("success", False):
                        # Fiziksel conversion başarılı - converted STEP dosyasını kullan
                        converted_step_path = physical_conversion_result["output_path"]
                        original_saved_path = physical_conversion_result["original_saved_path"]
                        
                        # Analysis için STEP dosya bilgisini oluştur
                        analysis_file_info = {
                            "original_filename": file_info["original_filename"],  # Orijinal adı koru
                            "saved_filename": os.path.basename(converted_step_path),
                            "file_path": converted_step_path,  # Converted STEP path
                            "file_size": os.path.getsize(converted_step_path) if os.path.exists(converted_step_path) else 0,
                            "file_type": "step",  # Artık STEP
                            "conversion_info": {
                                "needs_conversion": False,
                                "is_cad_file": True,
                                "original_format": file_info["conversion_info"]["original_format"],
                                "converted_from": file_info["conversion_info"]["original_format"],
                                "physical_conversion": True,
                                "original_file_path": file_info["file_path"],
                                "original_saved_path": original_saved_path,
                                "conversion_log_file": physical_conversion_result.get("log_file"),
                                "conversion_time": physical_conversion_result.get("processing_time", 0)
                            }
                        }
                        
                        cad_files.append(analysis_file_info)
                        
                        print(f"[MULTIPLE-UPLOAD-PHYSICAL] ✅ Fiziksel conversion başarılı: {file_info['original_filename']} -> {converted_step_path}")
                        
                    else:
                        # Conversion başarısız - detaylı hata bilgisi ile kaydet
                        error_msg = physical_conversion_result.get('error', 'Unknown error')
                        
                        # Extract format info from the conversion result
                        detected_format = physical_conversion_result.get('detected_format', 'Unknown')
                        format_confidence = physical_conversion_result.get('format_confidence', 0)
                        recommendations = physical_conversion_result.get('recommendations', [])
                        
                        # If format not in direct result, try to extract from error message
                        if detected_format == 'Unknown' and 'NX/Unigraphics' in error_msg:
                            detected_format = 'NX/Unigraphics'
                            format_confidence = 90
                        nx_help = physical_conversion_result.get('nx_help')
                        support_contact = physical_conversion_result.get('support_contact')
                        
                        # Build user-friendly error message
                        if detected_format == "NX/Unigraphics":
                            user_error = "Bu dosya NX CAD yazılımında oluşturulmuş. Otomatik dönüştürme için lisanslı NX yazılımı gerekli."
                            
                            # Get NX help if available
                            nx_help = physical_conversion_result.get('nx_help')
                            support_contact = physical_conversion_result.get('support_contact')
                            
                            # Add default recommendations for NX
                            if not recommendations or len(recommendations) == 0:
                                recommendations = [
                                    "NX yazılımında: File → Export → STEP 214",
                                    "CAD desteğe başvurun: cad-support@company.com", 
                                    "Alternatif: Teknik çizim PDF'ini yükleyin"
                                ]
                            
                            # Add NX-specific immediate solutions
                            immediate_solutions = []
                            if nx_help:
                                manual_steps = nx_help.get('conversion_options', {}).get('manual_conversion', {}).get('NX_Native', [])
                                if manual_steps:
                                    immediate_solutions.append({
                                        "title": "NX'den Export",
                                        "steps": manual_steps,
                                        "estimated_time": "5 dakika"
                                    })
                                else:
                                    # Default NX steps
                                    immediate_solutions.append({
                                        "title": "NX Yazılımı ile Dönüştürme",
                                        "steps": [
                                            "1. NX yazılımını açın",
                                            "2. File → Open ile PRT dosyasını açın",
                                            "3. File → Export → STEP seçin",
                                            "4. Export Options'da STEP 214 seçin",
                                            "5. Export butonuna tıklayın"
                                        ],
                                        "estimated_time": "5 dakika"
                                    })
                                
                                # Add support option
                                if support_contact:
                                    immediate_solutions.append({
                                        "title": "CAD Destek Ekibi",
                                        "steps": [
                                            f"Dosyayı {support_contact.get('email', 'cad-support@company.com')} adresine gönderin",
                                            "Konu: PRT to STEP Dönüştürme Talebi",
                                            f"Dosya adı: {file_info['original_filename']}",
                                            "24 saat içinde dönüştürülmüş dosyayı alacaksınız"
                                        ],
                                        "estimated_time": "1 iş günü"
                                    })
                            else:
                                # Default solutions
                                immediate_solutions = [
                                    {
                                        "title": "NX Yazılımı ile Dönüştürme",
                                        "steps": [
                                            "1. NX yazılımını açın",
                                            "2. File → Open ile PRT dosyasını açın",
                                            "3. File → Export → STEP seçin",
                                            "4. STEP 214 formatını seçin",
                                            "5. Export yapın"
                                        ],
                                        "estimated_time": "5 dakika"
                                    },
                                    {
                                        "title": "CAD Destek Ekibi",
                                        "steps": [
                                            "Dosyayı cad-support@company.com adresine gönderin",
                                            "Konu: PRT to STEP Dönüştürme",
                                            f"Dosya: {file_info['original_filename']}",
                                            "1 iş günü içinde dönüş alacaksınız"
                                        ],
                                        "estimated_time": "1 iş günü"
                                    }
                                ]
                        
                        elif detected_format == "Creo/Pro-E":
                            user_error = "Creo/Pro-E formatı desteklenmiyor. Lütfen Creo'dan STEP olarak kaydedin."
                        elif detected_format == "STEP":
                            user_error = "Dosya aslında STEP formatında. Uzantıyı .step olarak değiştirip tekrar yükleyin."
                        elif detected_format == "IGES":
                            user_error = "Dosya aslında IGES formatında. Uzantıyı .iges olarak değiştirip tekrar yükleyin."
                        elif detected_format == "Parasolid":
                            user_error = "Parasolid formatı sınırlı desteğe sahip. STEP formatında export önerilir."
                        else:
                            user_error = "PRT dosya formatı tanımlanamadı. Lütfen STEP formatında yeniden export yapın."
                        
                        # Create failed upload info
                        failed_upload_info = {
                            "filename": file_info['original_filename'],
                            "error": user_error,
                            "format_detected": detected_format,
                            "format_confidence": format_confidence,
                            "technical_details": {
                                "original_error": error_msg,
                                "detected_format": detected_format,
                                "format_confidence": format_confidence,
                                "file_size": file_info.get('file_size', 0),
                                "file_size_mb": round(file_info.get('file_size', 0) / (1024 * 1024), 2)
                            },
                            "recommendations": recommendations
                        }
                        
                        # Add NX-specific fields if available
                        if detected_format == "NX/Unigraphics" and nx_help:
                            failed_upload_info["immediate_solutions"] = immediate_solutions if 'immediate_solutions' in locals() else []
                            failed_upload_info["alternative_workflow"] = nx_help.get('alternative_workflow', {})
                            failed_upload_info["support_contact"] = support_contact
                            failed_upload_info["help_resources"] = {
                                "conversion_options": nx_help.get('conversion_options', {}),
                                "format_info": nx_help.get('format_info', {})
                            }
                        
                        failed_uploads.append(failed_upload_info)
                        
                        print(f"[MULTIPLE-UPLOAD-PHYSICAL] ❌ Fiziksel conversion başarısız: {file_info['original_filename']}")
                        print(f"[MULTIPLE-UPLOAD-PHYSICAL] 📋 Detected format: {detected_format} (confidence: {format_confidence}%)")
                        print(f"[MULTIPLE-UPLOAD-PHYSICAL] 💡 User message: {user_error}")
                        
                        if detected_format == "NX/Unigraphics" and support_contact:
                            print(f"[MULTIPLE-UPLOAD-PHYSICAL] 📧 Support contact: {support_contact.get('email', 'N/A')}")
                        
                        continue# Update for file_upload_controller.py - PRT conversion error handling
                
                else:
                    # Conversion gerekmez - direkt analiz için hazırla
                    analysis_file_info = file_info
                    analysis_file_info["conversion_info"]["physical_conversion"] = False
                    
                    conversion_results.append({
                        "success": True,
                        "conversion_needed": False,
                        "message": "No conversion needed"
                    })
                
                # Türe göre kategorilere ayır
                file_type = analysis_file_info['file_type']
                if file_type == 'pdf':
                    pdf_files.append(analysis_file_info)
                elif file_type in ['step', 'cad_part']:  # ✅ CAD dosyaları (converted dahil)
                    cad_files.append(analysis_file_info)
                else:
                    other_files.append(analysis_file_info)
                
                original_format = file_info.get('conversion_info', {}).get('original_format', file_type)
                conversion_info = ""
                if analysis_file_info.get('conversion_info', {}).get('physical_conversion', False):
                    conversion_info = f" (physically converted from {original_format.upper()})"
                elif file_type != original_format:
                    conversion_info = f" (memory converted from {original_format.upper()})"
                
                print(f"[MULTIPLE-UPLOAD-PHYSICAL] ✅ Kaydedildi: {file_info['original_filename']} ({file_type}){conversion_info}")
                
            except Exception as e:
                failed_uploads.append({
                    "filename": file.filename,
                    "error": str(e)
                })
                print(f"[MULTIPLE-UPLOAD-PHYSICAL] ❌ Dosya işleme hatası: {file.filename} - {str(e)}")
        
        print(f"[MULTIPLE-UPLOAD-PHYSICAL] 📊 PDF: {len(pdf_files)}, CAD: {len(cad_files)}, Other: {len(other_files)}, Failed: {len(failed_uploads)}")
        
        # ✅ ENHANCED - PDF-CAD EŞLEŞTİRME (Fiziksel conversion desteği)
        matched_pairs = []
        unmatched_pdfs = []
        unmatched_cads = []
        
        if pdf_files and cad_files:
            print(f"[MULTIPLE-UPLOAD-PHYSICAL] 🔄 PDF-CAD eşleştirme başlıyor (fiziksel dosyalar dahil)...")
            
            # Eşleştirmeleri hesapla
            matches = match_pdf_to_cad_files(pdf_files, cad_files)
            
            used_cad_files = set()
            
            for match in matches:
                pdf_info = match['pdf_file']
                cad_info = match['cad_file']
                score = match['match_score']
                
                # Minimum eşleştirme skoru kontrolü
                if cad_info and score >= 30.0 and cad_info['saved_filename'] not in used_cad_files:
                    matched_pairs.append(match)
                    used_cad_files.add(cad_info['saved_filename'])
                    
                    cad_format = cad_info.get('conversion_info', {}).get('original_format', 'STEP').upper()
                    conversion_type = "Physical" if cad_info.get('conversion_info', {}).get('physical_conversion', False) else "Direct"
                    print(f"[MULTIPLE-UPLOAD-PHYSICAL] ✅ Eşleşti: {pdf_info['original_filename']} ↔ {cad_info['original_filename']} ({cad_format}-{conversion_type}) ({score}%)")
                else:
                    unmatched_pdfs.append(pdf_info)
                    if score < 30.0:
                        print(f"[MULTIPLE-UPLOAD-PHYSICAL] ❌ Düşük skor: {pdf_info['original_filename']} ({score}%)")
            
            # Kullanılmayan CAD dosyalarını bul
            for cad_info in cad_files:
                if cad_info['saved_filename'] not in used_cad_files:
                    unmatched_cads.append(cad_info)
        else:
            # Eşleştirme yapılamaz
            unmatched_pdfs = pdf_files
            unmatched_cads = cad_files
        
        # ✅ ENHANCED - ANALİZ KAYITLARI OLUŞTUR (Fiziksel conversion bilgisi ile)
        created_analyses = []
        
        # 1. Eşleşmiş çiftler için
        for match in matched_pairs:
            pdf_info = match['pdf_file']
            cad_info = match['cad_file']
            
            # PDF analizi oluştur (Fiziksel CAD conversion bilgisi ile)
            analysis_data = {
                "user_id": current_user['id'],
                "filename": pdf_info['saved_filename'],
                "original_filename": pdf_info['original_filename'],
                "file_type": pdf_info['file_type'],
                "file_size": pdf_info['file_size'],
                "file_path": pdf_info['file_path'],
                "analysis_status": "uploaded",
                # ✅ ENHANCED - PDF-CAD eşleştirme bilgileri
                "matched_cad_file": cad_info['saved_filename'],
                "matched_cad_path": cad_info['file_path'],
                "match_score": match['match_score'],
                "match_quality": match['match_quality'],
                "analysis_strategy": match['analysis_strategy']
            }
            
            # ✅ YENİ - Fiziksel CAD conversion bilgilerini ekle
            cad_conversion_info = cad_info.get('conversion_info', {})
            if cad_conversion_info.get('physical_conversion', False):
                analysis_data.update({
                    "matched_cad_original_format": cad_conversion_info['original_format'],
                    "matched_cad_converted": True,
                    "matched_cad_physical_conversion": True,
                    "matched_cad_conversion_time": cad_conversion_info.get('conversion_time', 0),
                    "matched_cad_original_file_path": cad_conversion_info.get('original_file_path'),
                    "matched_cad_original_saved_path": cad_conversion_info.get('original_saved_path'),
                    "matched_cad_conversion_log": cad_conversion_info.get('conversion_log_file')
                })
            
            pdf_analysis = FileAnalysis.create_analysis(analysis_data)
            
            created_analyses.append({
                "analysis_id": pdf_analysis['id'],
                "type": "pdf_with_cad",  # ✅ ENHANCED
                "primary_file": pdf_info['original_filename'],
                "secondary_file": cad_info['original_filename'],
                "match_score": match['match_score'],
                "cad_physically_converted": cad_conversion_info.get('physical_conversion', False),
                "file_info": pdf_analysis
            })
        
        # 2. Eşleşmemiş PDF'ler için
        for pdf_info in unmatched_pdfs:
            pdf_analysis = FileAnalysis.create_analysis({
                "user_id": current_user['id'],
                "filename": pdf_info['saved_filename'],
                "original_filename": pdf_info['original_filename'],
                "file_type": pdf_info['file_type'],
                "file_size": pdf_info['file_size'],
                "file_path": pdf_info['file_path'],
                "analysis_status": "uploaded",
                "analysis_strategy": "pdf_only_extract_step"
            })
            
            created_analyses.append({
                "analysis_id": pdf_analysis['id'],
                "type": "pdf_only",
                "primary_file": pdf_info['original_filename'],
                "file_info": pdf_analysis
            })
        
        # 3. Eşleşmemiş CAD'ler için
        for cad_info in unmatched_cads:
            analysis_data = {
                "user_id": current_user['id'],
                "filename": cad_info['saved_filename'],
                "original_filename": cad_info['original_filename'],
                "file_type": cad_info['file_type'],
                "file_size": cad_info['file_size'],
                "file_path": cad_info['file_path'],
                "analysis_status": "uploaded"
            }
            
            # ✅ YENİ - Fiziksel CAD conversion bilgilerini ekle
            cad_conversion_info = cad_info.get('conversion_info', {})
            if cad_conversion_info.get('physical_conversion', False):
                analysis_data.update({
                    "cad_converted": True,
                    "cad_physical_conversion": True,
                    "original_cad_format": cad_conversion_info['original_format'],
                    "conversion_time": cad_conversion_info.get('conversion_time', 0),
                    "original_file_path": cad_conversion_info.get('original_file_path'),
                    "original_saved_path": cad_conversion_info.get('original_saved_path'),
                    "conversion_log_file": cad_conversion_info.get('conversion_log_file')
                })
            
            cad_analysis = FileAnalysis.create_analysis(analysis_data)
            
            cad_type = "cad_only"
            if cad_info['file_type'] == 'step':
                if cad_conversion_info.get('physical_conversion', False):
                    cad_type = "step_converted"
                else:
                    cad_type = "step_only"
            
            created_analyses.append({
                "analysis_id": cad_analysis['id'],
                "type": cad_type,  # ✅ ENHANCED
                "primary_file": cad_info['original_filename'],
                "physically_converted": cad_conversion_info.get('physical_conversion', False),
                "file_info": cad_analysis
            })
        
        # 4. Diğer dosyalar için
        for other_info in other_files:
            other_analysis = FileAnalysis.create_analysis({
                "user_id": current_user['id'],
                "filename": other_info['saved_filename'],
                "original_filename": other_info['original_filename'],
                "file_type": other_info['file_type'],
                "file_size": other_info['file_size'],
                "file_path": other_info['file_path'],
                "analysis_status": "uploaded"
            })
            
            created_analyses.append({
                "analysis_id": other_analysis['id'],
                "type": "document",
                "primary_file": other_info['original_filename'],
                "file_info": other_analysis
            })
        
        # ✅ ENHANCED - SONUÇ HAZIRLA (Fiziksel CAD conversion desteği bilgileri ile)
        successful_physical_conversions = len([r for r in conversion_results if r.get("success", False) and r.get("conversion_needed", False)])
        failed_physical_conversions = len([r for r in conversion_results if not r.get("success", False) and r.get("conversion_needed", False)])
        
        response_data = {
            "success": True,
            "message": f"{len(created_analyses)} dosya başarıyla yüklendi ve analiz için hazırlandı (Fiziksel CAD conversion aktif)",
            "upload_summary": {
                "total_uploaded": len(created_analyses),
                "pdf_files": len(pdf_files),
                "cad_files": len(cad_files),  # ✅ STEP + Converted PRT + Converted CATPART
                "other_files": len(other_files),
                "failed_uploads": len(failed_uploads),
                "matched_pairs": len(matched_pairs),
                "unmatched_pdfs": len(unmatched_pdfs),
                "unmatched_cads": len(unmatched_cads),  # ✅ CAD
                # ✅ YENİ - Fiziksel Conversion istatistikleri
                "physical_cad_conversions": {
                    "total_attempted": len([r for r in conversion_results if r.get("conversion_needed", False)]),
                    "successful": successful_physical_conversions,
                    "failed": failed_physical_conversions,
                    "conversion_engine": "CADConverterService",
                    "output_directory": cad_converter.step_files_dir,
                    "original_files_directory": cad_converter.original_files_dir
                }
            },
            "analyses": created_analyses,
            "matching_results": {
                "pdf_cad_matches": [  # ✅ CAD
                    {
                        "pdf_file": match['pdf_file']['original_filename'],
                        "cad_file": match['cad_file']['original_filename'],
                        "cad_format": match['cad_file'].get('conversion_info', {}).get('original_format', 'STEP').upper(),
                        "physically_converted": match['cad_file'].get('conversion_info', {}).get('physical_conversion', False),
                        "match_score": match['match_score'],
                        "match_quality": match['match_quality']
                    }
                    for match in matched_pairs
                ],
                "unmatched_files": {
                    "pdfs": [f['original_filename'] for f in unmatched_pdfs],
                    "cads": [f['original_filename'] for f in unmatched_cads]  # ✅ CAD
                }
            },
            "failed_uploads": failed_uploads,
            "physical_conversion_results": physical_conversion_log,  # ✅ YENİ - Detaylı conversion log
            "physical_conversion_summary": {  # ✅ YENİ - Conversion özeti
                "total_attempted": len([log for log in physical_conversion_log if log.get("conversion_successful") is not None]),
                "successful": len([log for log in physical_conversion_log if log.get("conversion_successful", False)]),
                "failed": len([log for log in physical_conversion_log if not log.get("conversion_successful", True)]),
                "total_processing_time": sum([log.get("processing_time", 0) for log in physical_conversion_log]),
                "output_directory_info": cad_converter.get_output_directory_info()
            },
            "next_steps": {
                "analyze_all": f"/api/upload/batch-analyze",
                "analyze_individual": f"/api/upload/analyze/{{analysis_id}}",
                "view_conversion_results": f"/api/upload/conversion-directory-info"
            }
        }
        
        print(f"[MULTIPLE-UPLOAD-PHYSICAL] ✅ Tamamlandı: {len(created_analyses)} analiz oluşturuldu")
        print(f"[MULTIPLE-UPLOAD-PHYSICAL] 🔧 Fiziksel conversions: {successful_physical_conversions} başarılı, {failed_physical_conversions} başarısız")
        print(f"[MULTIPLE-UPLOAD-PHYSICAL] 📁 Output directory: {cad_converter.step_files_dir}")
        
        return jsonify(response_data), 201
        
    except Exception as e:
        print(f"[MULTIPLE-UPLOAD-PHYSICAL] ❌ Hata: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            "success": False,
            "message": f"Çoklu dosya yükleme hatası: {str(e)}"
        }), 500


# ===== YENİ ENDPOINT - CONVERSION DIRECTORY INFO =====

@upload_bp.route('/conversion-directory-info', methods=['GET'])
@jwt_required()
def get_conversion_directory_info():
    """Fiziksel CAD conversion output directory bilgileri"""
    try:
        current_user = get_current_user()
        
        # Directory bilgilerini al
        dir_info = cad_converter.get_output_directory_info()
        
        # CAD converter status
        converter_status = cad_converter.get_status()
        
        return jsonify({
            "success": True,
            "conversion_directories": dir_info,
            "converter_status": converter_status,
            "access_info": {
                "step_files_url": f"/static/cad_conversions/",  # Web erişim için
                "download_endpoint": "/api/upload/download-converted-file/"
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Directory info error: {str(e)}"
        }), 500

@upload_bp.route('/download-converted-file/<path:file_path>', methods=['GET'])
@jwt_required()
def download_converted_file(file_path):
    """Fiziksel olarak convert edilmiş dosyayı indir"""
    try:
        current_user = get_current_user()
        
        # Güvenlik: sadece conversion directory altındaki dosyalara erişim
        if not file_path.startswith(cad_converter.output_base_dir):
            full_path = os.path.join(cad_converter.output_base_dir, file_path)
        else:
            full_path = file_path
        
        # Path traversal güvenliği
        full_path = os.path.abspath(full_path)
        base_path = os.path.abspath(cad_converter.output_base_dir)
        
        if not full_path.startswith(base_path):
            return jsonify({
                "success": False,
                "message": "Geçersiz dosya yolu"
            }), 403
        
        if not os.path.exists(full_path):
            return jsonify({
                "success": False,
                "message": "Dosya bulunamadı"
            }), 404
        
        # Dosyayı indir
        return send_file(
            full_path,
            as_attachment=True,
            download_name=os.path.basename(full_path)
        )
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Download error: {str(e)}"
        }), 500

@upload_bp.route('/analyze/<analysis_id>', methods=['POST'])
@jwt_required()
def analyze_uploaded_file_enhanced(analysis_id):
    """✅ ENHANCED - Analiz + PDF-STEP eşleştirme desteği + FIXED CAD FILE HANDLING"""
    try:
        current_user = get_current_user()
        
        # ✅ 1. FAST VALIDATION
        analysis = FileAnalysis.find_by_id(analysis_id)
        if not analysis:
            return jsonify({"success": False, "message": "Analiz kaydı bulunamadı"}), 404
        
        if analysis['user_id'] != current_user['id']:
            return jsonify({"success": False, "message": "Bu dosyaya erişim yetkiniz yok"}), 403
        
        if analysis['analysis_status'] == 'analyzing':
            return jsonify({"success": False, "message": "Dosya zaten analiz ediliyor"}), 409
        
        # ✅ 2. ENHANCED FILE PATH DETECTION - CAD CONVERSION DESTEĞİ
        analysis_file_path = None
        file_source = "original"
        
        print(f"[ANALYZE-ENHANCED] 🔍 Dosya yolu tespiti başlıyor: {analysis.get('original_filename')}")
        print(f"[ANALYZE-ENHANCED] 📋 Database paths:")
        print(f"   - file_path: {analysis.get('file_path')}")
        print(f"   - converted_step_path: {analysis.get('converted_step_path')}")
        print(f"   - matched_cad_path: {analysis.get('matched_cad_path')}")
        print(f"   - matched_step_path: {analysis.get('matched_step_path')}")
        
        # Öncelik sırası: converted_step_path > matched_cad_path > matched_step_path > original_file_path
        if analysis.get('converted_step_path') and os.path.exists(analysis['converted_step_path']):
            analysis_file_path = analysis['converted_step_path']
            file_source = "converted_step"
            print(f"[ANALYZE-ENHANCED] ✅ Using converted STEP: {analysis_file_path}")
            
        elif analysis.get('matched_cad_path') and os.path.exists(analysis['matched_cad_path']):
            analysis_file_path = analysis['matched_cad_path']
            file_source = "matched_cad"
            print(f"[ANALYZE-ENHANCED] ✅ Using matched CAD: {analysis_file_path}")
            
        elif analysis.get('matched_step_path') and os.path.exists(analysis['matched_step_path']):
            analysis_file_path = analysis['matched_step_path']
            file_source = "matched_step"
            print(f"[ANALYZE-ENHANCED] ✅ Using matched STEP: {analysis_file_path}")
            
        elif analysis.get('file_path') and os.path.exists(analysis['file_path']):
            analysis_file_path = analysis['file_path']
            file_source = "original"
            print(f"[ANALYZE-ENHANCED] ✅ Using original file: {analysis_file_path}")
            
        else:
            # ✅ FALLBACK: Tüm olası konumları kontrol et
            print(f"[ANALYZE-ENHANCED] 🔍 Database paths bulunamadı, fallback araması başlıyor...")
            
            possible_paths = []
            original_filename = analysis.get('original_filename', '')
            
            # 1. Database'deki tüm paths'leri ekle (var olmasalar bile)
            for path_key in ['file_path', 'converted_step_path', 'matched_cad_path', 'matched_step_path']:
                path_value = analysis.get(path_key)
                if path_value:
                    possible_paths.append(path_value)
            
            # 2. CAD conversion klasöründeki olası dosyalar
            if original_filename:
                base_name = os.path.splitext(original_filename)[0]
                conversion_dir = "/app/temp/cad_conversion"
                
                print(f"[ANALYZE-ENHANCED] 🔍 Conversion klasörü kontrol ediliyor: {conversion_dir}")
                
                if os.path.exists(conversion_dir):
                    try:
                        for file in os.listdir(conversion_dir):
                            # Base name veya tam filename ile eşleşen STEP dosyaları
                            if (base_name.lower() in file.lower() or 
                                original_filename.lower().replace('.prt', '').replace('.catpart', '') in file.lower()) and \
                               file.endswith('.step'):
                                possible_paths.append(os.path.join(conversion_dir, file))
                                print(f"[ANALYZE-ENHANCED] 🔍 Conversion'da bulundu: {file}")
                    except Exception as e:
                        print(f"[ANALYZE-ENHANCED] ⚠️ Conversion klasör tarama hatası: {e}")
            
            # 3. Upload klasöründeki dosyalar
            upload_dir = "/app/uploads"
            if os.path.exists(upload_dir) and original_filename:
                print(f"[ANALYZE-ENHANCED] 🔍 Upload klasörü kontrol ediliyor: {upload_dir}")
                
                try:
                    for file in os.listdir(upload_dir):
                        # Orijinal dosya adını içeren dosyalar
                        if original_filename.lower() in file.lower():
                            possible_paths.append(os.path.join(upload_dir, file))
                            print(f"[ANALYZE-ENHANCED] 🔍 Upload'da bulundu: {file}")
                except Exception as e:
                    print(f"[ANALYZE-ENHANCED] ⚠️ Upload klasör tarama hatası: {e}")
            
            # 4. Analysis ID bazlı static klasör kontrol et
            static_step_dir = f"/app/static/stepviews/{analysis_id}"
            if os.path.exists(static_step_dir):
                try:
                    for file in os.listdir(static_step_dir):
                        if file.endswith(('.step', '.stp')):
                            possible_paths.append(os.path.join(static_step_dir, file))
                            print(f"[ANALYZE-ENHANCED] 🔍 Static'te bulundu: {file}")
                except Exception as e:
                    print(f"[ANALYZE-ENHANCED] ⚠️ Static klasör tarama hatası: {e}")
            
            print(f"[ANALYZE-ENHANCED] 📋 Toplam {len(possible_paths)} olası path bulundu")
            
            # İlk var olan dosyayı kullan (STEP dosyalarına öncelik ver)
            step_files = [p for p in possible_paths if p.endswith(('.step', '.stp'))]
            other_files = [p for p in possible_paths if not p.endswith(('.step', '.stp'))]
            
            # STEP dosyaları önce dene
            for path in step_files + other_files:
                if os.path.exists(path):
                    analysis_file_path = path
                    file_source = "fallback_found"
                    print(f"[ANALYZE-ENHANCED] ✅ Fallback ile bulundu: {analysis_file_path}")
                    break
        
        # Hala dosya bulunamadıysa detaylı hata ver
        if not analysis_file_path or not os.path.exists(analysis_file_path):
            error_details = {
                "analysis_id": analysis_id,
                "original_filename": analysis.get('original_filename'),
                "file_type": analysis.get('file_type'),
                "database_paths": {
                    "file_path": analysis.get('file_path'),
                    "converted_step_path": analysis.get('converted_step_path'),
                    "matched_cad_path": analysis.get('matched_cad_path'),
                    "matched_step_path": analysis.get('matched_step_path')
                },
                "checked_paths": possible_paths if 'possible_paths' in locals() else [],
                "directories_checked": [
                    "/app/uploads",
                    "/app/temp/cad_conversion",
                    f"/app/static/stepviews/{analysis_id}"
                ]
            }
            
            print(f"[ANALYZE-ENHANCED] ❌ File not found anywhere:")
            print(f"   Original filename: {analysis.get('original_filename')}")
            print(f"   File type: {analysis.get('file_type')}")
            print(f"   Checked paths: {error_details['checked_paths']}")
            
            return jsonify({
                "success": False, 
                "message": "Dosya sistemde bulunamadı",
                "error_code": "FILE_NOT_FOUND",
                "debug_info": error_details
            }), 404
        
        # ✅ 3. IMMEDIATE STATUS UPDATE WITH CORRECT FILE PATH
        FileAnalysis.update_analysis(analysis_id, {
            "analysis_status": "analyzing",
            "processing_time": None,
            "error_message": None,
            "actual_file_path": analysis_file_path,  # Gerçek kullanılan dosya yolu
            "file_source": file_source  # Dosya kaynağı
        })
        
        print(f"[ANALYZE-ENHANCED] ⚡ Gelişmiş analiz başlatılıyor: {analysis['original_filename']}")
        print(f"[ANALYZE-ENHANCED] 📂 File source: {file_source}")
        print(f"[ANALYZE-ENHANCED] 📁 File path: {analysis_file_path}")
        start_time = time.time()
        
        # ✅ 4. ENHANCED ANALYSIS WITH PDF-STEP MATCHING
        try:
            material_service = MaterialAnalysisService()
            
            # Eşleşmiş STEP dosyası var mı kontrol et
            matched_step_path = analysis.get('matched_step_path') or analysis.get('matched_cad_path')
            analysis_strategy = analysis.get('analysis_strategy', 'default')
            
            print(f"[ANALYZE-ENHANCED] 📋 Analiz stratejisi: {analysis_strategy}")
            if matched_step_path:
                print(f"[ANALYZE-ENHANCED] 🔗 Eşleşmiş STEP: {matched_step_path}")
            
            # ✅ ANALYSIS: Doğru dosya yolu ve dosya tipi ile analiz et
            # Dosya tipini dosya yolundan belirle
            actual_file_type = analysis.get('file_type', 'unknown')
            if analysis_file_path.endswith(('.step', '.stp')):
                actual_file_type = 'step'
            elif analysis_file_path.endswith('.pdf'):
                actual_file_type = 'pdf'
            elif analysis_file_path.endswith(('.doc', '.docx')):
                actual_file_type = 'document'
            
            print(f"[ANALYZE-ENHANCED] 🔧 Analysis parameters:")
            print(f"   - File path: {analysis_file_path}")
            print(f"   - File type: {actual_file_type}")
            print(f"   - User ID: {current_user['id']}")
            
            result = material_service.analyze_document_ultra_fast(
                analysis_file_path,  # ✅ Doğru dosya yolu
                actual_file_type,    # ✅ Doğru dosya tipi
                current_user['id']
            )
            
            # ✅ ENHANCED POST-PROCESSING
            # Eğer eşleşmiş STEP dosyası varsa ve PDF'den STEP çıkarılamadıysa, 
            # eşleşmiş STEP'i kullan
            if matched_step_path and os.path.exists(matched_step_path) and actual_file_type == 'pdf':
                if not result.get('step_analysis') or not result.get('step_file_hash'):
                    print(f"[ANALYZE-ENHANCED] 🔄 PDF'den STEP çıkarılamadı, eşleşmiş STEP kullanılıyor: {matched_step_path}")
                    
                    # Eşleşmiş STEP dosyasını analiz et
                    try:
                        import cadquery as cq
                        
                        assembly = cq.importers.importStep(matched_step_path)
                        shapes = assembly.objects
                        sorted_shapes = sorted(shapes, key=lambda s: s.Volume(), reverse=True)
                        main_shape = sorted_shapes[0]
                        main_bbox = main_shape.BoundingBox()
                        
                        relevant_shapes = [main_shape]
                        for shape in sorted_shapes[1:]:
                            bb = shape.BoundingBox()
                            intersects = (
                                bb.xmax > main_bbox.xmin and bb.xmin < main_bbox.xmax and
                                bb.ymax > main_bbox.ymin and bb.ymin < main_bbox.ymax and
                                bb.zmax > main_bbox.zmin and bb.zmin < main_bbox.zmax
                            )
                            if intersects:
                                relevant_shapes.append(shape)
                        
                        part = cq.Compound.makeCompound(relevant_shapes)
                        
                        # Boyut optimizasyonu
                        min_volume = None
                        best_dims = (0, 0, 0)
                        for rx in [0, 90, 180, 270]:
                            for ry in [0, 90, 180, 270]:
                                for rz in [0, 90, 180, 270]:
                                    rotated = part.rotate((0, 0, 0), (1, 0, 0), rx)\
                                                  .rotate((0, 0, 0), (0, 1, 0), ry)\
                                                  .rotate((0, 0, 0), (0, 0, 1), rz)
                                    bbox = rotated.BoundingBox()
                                    volume = bbox.xlen * bbox.ylen * bbox.zlen
                                    if (min_volume is None) or (volume < min_volume):
                                        min_volume = volume
                                        best_dims = (bbox.xlen, bbox.ylen, bbox.zlen)
                        
                        x, y, z = best_dims
                        
                        def always_round_up(value):
                            return int(value) if abs(value - int(value)) < 0.01 else int(value) + 1
                        
                        x_pad = always_round_up(x + 10.0)
                        y_pad = always_round_up(y + 10.0)
                        z_pad = always_round_up(z + 10.0)
                        volume_padded = x_pad * y_pad * z_pad
                        product_volume = part.Volume()
                        waste_volume = volume_padded - product_volume
                        waste_ratio = (waste_volume / volume_padded * 100) if volume_padded > 0 else 0.0
                        total_surface_area = part.Area()
                        
                        # Eşleşmiş STEP analiz sonucunu result'a ekle
                        result['step_analysis'] = {
                            "X (mm)": round(x, 3),
                            "Y (mm)": round(y, 3),
                            "Z (mm)": round(z, 3),
                            "Silindirik Çap (mm)": round(max(x, y), 3),
                            "Silindirik Yükseklik (mm)": round(z, 3),
                            "X+Pad (mm)": round(x_pad, 3),
                            "Y+Pad (mm)": round(y_pad, 3),
                            "Z+Pad (mm)": round(z_pad, 3),
                            "Prizma Hacmi (mm³)": round(volume_padded, 3),
                            "Ürün Hacmi (mm³)": round(product_volume, 3),
                            "Talaş Hacmi (mm³)": round(waste_volume, 3),
                            "Talaş Oranı (%)": round(waste_ratio, 2),
                            "Toplam Yüzey Alanı (mm²)": round(total_surface_area, 3)
                        }
                        
                        result['step_source'] = 'matched'
                        result['matched_step_used'] = True
                        
                        print(f"[ANALYZE-ENHANCED] ✅ Eşleşmiş STEP analizi tamamlandı: {matched_step_path}")
                        
                    except Exception as matched_step_error:
                        print(f"[ANALYZE-ENHANCED] ❌ Eşleşmiş STEP analiz hatası: {matched_step_error}")
                        result['step_source'] = 'none'
                        result['matched_step_used'] = False
                else:
                    print(f"[ANALYZE-ENHANCED] ✅ PDF'den STEP çıkarıldı, eşleşmiş STEP'e gerek yok")
                    result['step_source'] = 'extracted'
                    result['matched_step_used'] = False
            else:
                if result.get('step_analysis'):
                    result['step_source'] = 'extracted'
                else:
                    result['step_source'] = 'none'
                result['matched_step_used'] = False
            
            processing_time = time.time() - start_time
            print(f"[ANALYZE-ENHANCED] ⏱️ Core analiz tamamlandı: {processing_time:.2f}s")
            
            if not result.get('error'):
                # ✅ 5. INSTANT DATABASE UPDATE WITH ENHANCED DATA
                update_data = {
                    "analysis_status": "completed",
                    "processing_time": processing_time,
                    "material_matches": result.get('material_matches', []),
                    "best_material_block": result.get('best_block', ''),
                    "step_analysis": result.get('step_analysis', {}),
                    "cost_estimation": result.get('cost_estimation', {}),
                    "ai_price_prediction": result.get('ai_price_prediction', {}),
                    "all_material_calculations": result.get('all_material_calculations', []),
                    "material_options": result.get('material_options', []),
                    "processing_log": result.get('processing_log', []),
                    # Enhanced fields
                    "used_matched_step": bool(matched_step_path and result.get('matched_step_used', False)),
                    "step_source": result.get('step_source', 'none'),  # 'matched', 'extracted', 'none'
                    "material_confidence": result.get('material_confidence', 0),
                    # File handling details
                    "final_file_path": analysis_file_path,
                    "final_file_source": file_source,
                    "final_file_type": actual_file_type,
                    # Render fields
                    "render_status": "pending",
                    "enhanced_renders": {},
                    "isometric_view": None,
                    "stl_generated": False
                }
                
                # PDF specific fields
                if actual_file_type == 'pdf':
                    update_data.update({
                        "pdf_step_extracted": bool(result.get('step_file_hash')),
                        "extracted_step_path": result.get('extracted_step_path'),
                        "step_file_hash": result.get('step_file_hash')
                    })
                
                FileAnalysis.update_analysis(analysis_id, update_data)
                
                # ✅ 6. BACKGROUND RENDERING DECISION
                should_render = False
                render_path = None
                
                # Rendering priority: matched_step > extracted_step > direct_step > converted_step
                if matched_step_path and os.path.exists(matched_step_path):
                    should_render = True
                    render_path = matched_step_path
                    print(f"[ANALYZE-ENHANCED] 🎨 Rendering: Matched STEP - {matched_step_path}")
                elif actual_file_type in ['step', 'stp']:
                    should_render = True
                    render_path = analysis_file_path
                    print(f"[ANALYZE-ENHANCED] 🎨 Rendering: Direct STEP - {render_path}")
                elif result.get('extracted_step_path') and os.path.exists(result['extracted_step_path']):
                    should_render = True
                    render_path = result['extracted_step_path']
                    print(f"[ANALYZE-ENHANCED] 🎨 Rendering: Extracted STEP - {render_path}")
                elif file_source == "converted_step" and analysis_file_path.endswith('.step'):
                    should_render = True
                    render_path = analysis_file_path
                    print(f"[ANALYZE-ENHANCED] 🎨 Rendering: Converted STEP - {render_path}")
                
                if should_render and render_path:
                    # Queue background rendering task
                    task_id = bg_processor.add_task(
                        background_render_task_enhanced,
                        args=(analysis_id, render_path, analysis_strategy),
                        kwargs={}
                    )
                    
                    # Update render status
                    FileAnalysis.update_analysis(analysis_id, {
                        "render_task_id": task_id,
                        "render_status": "processing"
                    })
                    
                    print(f"[ANALYZE-ENHANCED] 🎨 Background render queued: {task_id}")
                
                # ✅ 7. ENHANCED INSTANT RESPONSE
                updated_analysis = FileAnalysis.find_by_id(analysis_id)
                
                response_data = {
                    "success": True,
                    "message": "Gelişmiş analiz başarıyla tamamlandı",
                    "analysis": updated_analysis,
                    "processing_time": processing_time,
                    "render_status": "processing" if should_render else "not_applicable",
                    "file_handling": {
                        "file_path_used": analysis_file_path,
                        "file_source": file_source,
                        "file_type_detected": actual_file_type,
                        "file_exists": os.path.exists(analysis_file_path)
                    },
                    "enhancement_details": {
                        "used_matched_step": bool(matched_step_path and result.get('matched_step_used', False)),
                        "step_source": result.get('step_source', 'none'),
                        "material_confidence": result.get('material_confidence', 0),
                        "analysis_strategy": analysis_strategy,
                        "match_score": analysis.get('match_score'),
                        "pdf_step_extracted": actual_file_type == 'pdf' and bool(result.get('step_file_hash'))
                    },
                    "analysis_details": {
                        "material_matches_count": len(result.get('material_matches', [])),
                        "step_analysis_available": bool(result.get('step_analysis')),
                        "cost_estimation_available": bool(result.get('cost_estimation')),
                        "material_calculations_count": len(result.get('all_material_calculations', [])),
                        "render_will_be_available": should_render,
                        "estimated_render_time": "30-60 seconds" if should_render else "N/A"
                    }
                }
                
                print(f"[ANALYZE-ENHANCED] 📤 Enhanced response sent: {processing_time:.2f}s")
                return jsonify(response_data), 200
            
            else:
                # Analysis error
                error_msg = result.get('error', 'Bilinmeyen analiz hatası')
                FileAnalysis.update_analysis(analysis_id, {
                    "analysis_status": "failed",
                    "error_message": error_msg,
                    "processing_time": time.time() - start_time,
                    "failed_file_path": analysis_file_path,
                    "failed_file_source": file_source
                })
                
                return jsonify({
                    "success": False,
                    "message": f"Analiz hatası: {error_msg}",
                    "file_info": {
                        "file_path": analysis_file_path,
                        "file_source": file_source
                    }
                }), 500
                
        except Exception as analysis_error:
            error_message = f"Analysis Service hatası: {str(analysis_error)}"
            FileAnalysis.update_analysis(analysis_id, {
                "analysis_status": "failed",
                "error_message": error_message,
                "processing_time": time.time() - start_time,
                "failed_file_path": analysis_file_path if 'analysis_file_path' in locals() else None,
                "failed_file_source": file_source if 'file_source' in locals() else None
            })
            
            print(f"[ANALYZE-ENHANCED] ❌ Analysis error: {error_message}")
            import traceback
            print(f"[ANALYZE-ENHANCED] 📋 Traceback: {traceback.format_exc()}")
            
            return jsonify({
                "success": False,
                "message": error_message,
                "file_info": {
                    "file_path": analysis_file_path if 'analysis_file_path' in locals() else None,
                    "file_source": file_source if 'file_source' in locals() else None
                }
            }), 500
        
    except Exception as e:
        try:
            FileAnalysis.update_analysis(analysis_id, {
                "analysis_status": "failed",
                "error_message": str(e)
            })
        except:
            pass
            
        print(f"[ANALYZE-ENHANCED] ❌ Unexpected error: {str(e)}")
        import traceback
        print(f"[ANALYZE-ENHANCED] 📋 Traceback: {traceback.format_exc()}")
        
        return jsonify({
            "success": False,
            "message": f"Beklenmeyen hata: {str(e)}"
        }), 500
# ===== RENDER ENDPOINTS =====

@upload_bp.route('/render/<analysis_id>', methods=['POST'])
@jwt_required()
def generate_step_render(analysis_id):
    """STEP dosyası için render oluştur"""
    try:
        current_user = get_current_user()
        
        # Analiz kaydını bul
        analysis = FileAnalysis.find_by_id(analysis_id)
        if not analysis:
            return jsonify({
                "success": False,
                "message": "Analiz kaydı bulunamadı"
            }), 404
        
        # Kullanıcı yetkisi kontrolü
        if analysis['user_id'] != current_user['id']:
            return jsonify({
                "success": False,
                "message": "Bu dosyaya erişim yetkiniz yok"
            }), 403
        
        # STEP dosyası kontrolü
        if analysis['file_type'] not in ['step', 'stp'] and not analysis.get('matched_step_path'):
            return jsonify({
                "success": False,
                "message": "Render için STEP dosyası gerekli"
            }), 400
        
        # Render parametrelerini al
        request_data = request.get_json() or {}
        include_dimensions = request_data.get('include_dimensions', True)
        include_materials = request_data.get('include_materials', True)
        high_quality = request_data.get('high_quality', True)
        
        print(f"[STEP-RENDER] 🎨 Render isteği: {analysis_id}")
        
        # STEP dosya yolunu belirle
        if analysis.get('matched_step_path'):
            step_path = analysis['matched_step_path']
        else:
            step_path = analysis['file_path']
        
        # Dosya varlık kontrolü
        if not os.path.exists(step_path):
            return jsonify({
                "success": False,
                "message": "STEP dosyası sistemde bulunamadı"
            }), 404
        
        # Step Renderer'ı kullan
        step_renderer = StepRendererEnhanced()
        
        render_result = step_renderer.generate_comprehensive_views(
            step_path,
            analysis_id=analysis_id,
            include_dimensions=include_dimensions,
            include_materials=include_materials,
            high_quality=high_quality
        )
        
        if render_result['success']:
            # Analiz kaydını güncelle
            update_data = {
                "enhanced_renders": render_result['renders'],
                "render_quality": "high" if high_quality else "standard",
                "render_status": "completed"
            }
            
            # Ana isometric view'ı ekle
            if 'isometric' in render_result['renders']:
                update_data["isometric_view"] = render_result['renders']['isometric']['file_path']
                if 'excel_path' in render_result['renders']['isometric']:
                    update_data["isometric_view_clean"] = render_result['renders']['isometric']['excel_path']
            
            FileAnalysis.update_analysis(analysis_id, update_data)
            
            return jsonify({
                "success": True,
                "message": "Render başarıyla oluşturuldu",
                "renders": render_result['renders'],
                "session_id": render_result['session_id'],
                "dimensions": render_result['dimensions'],
                "total_views": render_result['total_views']
            }), 200
        else:
            return jsonify({
                "success": False,
                "message": f"Render oluşturma başarısız: {render_result.get('message', 'Bilinmeyen hata')}"
            }), 500
            
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Render hatası: {str(e)}"
        }), 500

@upload_bp.route('/render-status/<analysis_id>', methods=['GET'])
@jwt_required()
def get_render_status_enhanced(analysis_id):
    """Enhanced render durumunu kontrol et - DEBUGGING"""
    try:
        current_user = get_current_user()
        
        analysis = FileAnalysis.find_by_id(analysis_id)
        if not analysis:
            return jsonify({
                "success": False,
                "message": "Analiz kaydı bulunamadı"
            }), 404
        
        # Kullanıcı yetkisi kontrolü
        if analysis['user_id'] != current_user['id']:
            return jsonify({
                "success": False,
                "message": "Bu dosyaya erişim yetkiniz yok"
            }), 403
        
        # Render durumunu kontrol et
        render_status = analysis.get('render_status', 'none')
        render_task_id = analysis.get('render_task_id')
        enhanced_renders = analysis.get('enhanced_renders', {})
        
        print(f"[RENDER-STATUS] 🔍 Analysis {analysis_id}:")
        print(f"   - render_status: {render_status}")
        print(f"   - render_task_id: {render_task_id}")
        print(f"   - enhanced_renders count: {len(enhanced_renders)}")
        print(f"   - enhanced_renders keys: {list(enhanced_renders.keys())}")
        
        response = {
            "success": True,
            "render_status": render_status,
            "has_renders": bool(enhanced_renders),
            "render_count": len(enhanced_renders),
            "render_details": enhanced_renders,  # Full details for debugging
            "stl_generated": analysis.get('stl_generated', False),
            "stl_path": analysis.get('stl_path'),
            "isometric_view": analysis.get('isometric_view'),
            "render_quality": analysis.get('render_quality', 'none'),
            "render_strategy": analysis.get('render_strategy'),
            "last_render_update": analysis.get('last_render_update'),
            "render_error": analysis.get('render_error'),
            # Debug fields
            "debug_info": {
                "analysis_id": analysis_id,
                "file_type": analysis.get('file_type'),
                "original_filename": analysis.get('original_filename'),
                "step_analysis_available": bool(analysis.get('step_analysis')),
                "extracted_step_path": analysis.get('extracted_step_path'),
                "matched_step_path": analysis.get('matched_step_path')
            }
        }
        
        # Background task durumunu kontrol et
        if render_task_id:
            task_result = bg_processor.get_result(render_task_id)
            if task_result:
                response["background_task"] = task_result
                print(f"[RENDER-STATUS] 🔄 Background task result: {task_result}")
            else:
                print(f"[RENDER-STATUS] ⏳ Background task still running: {render_task_id}")
        
        # Render'lar hazırsa detayları ekle
        if render_status == 'completed' and enhanced_renders:
            response["renders"] = {}
            for view_name, view_data in enhanced_renders.items():
                if view_data.get('success'):
                    response["renders"][view_name] = {
                        "file_path": view_data.get('file_path'),
                        "excel_path": view_data.get('excel_path'),
                        "file_exists": os.path.exists(os.path.join(os.getcwd(), view_data.get('file_path', '').lstrip('/'))) if view_data.get('file_path') else False
                    }
        
        return jsonify(response), 200
        
    except Exception as e:
        import traceback
        print(f"[RENDER-STATUS] ❌ Error: {str(e)}")
        print(f"[RENDER-STATUS] 📋 Traceback: {traceback.format_exc()}")
        
        return jsonify({
            "success": False,
            "message": f"Durum kontrolü hatası: {str(e)}"
        }), 500

# ===== STL GENERATION =====

@upload_bp.route('/generate-stl/<analysis_id>', methods=['POST'])
@jwt_required()
def generate_stl_for_analysis(analysis_id):
    """Analiz için STL dosyası oluştur"""
    try:
        current_user = get_current_user()
        
        # Analiz kaydını bul
        analysis = FileAnalysis.find_by_id(analysis_id)
        if not analysis:
            return jsonify({
                "success": False,
                "message": "Analiz kaydı bulunamadı"
            }), 404
        
        # Kullanıcı yetkisi kontrolü
        if analysis['user_id'] != current_user['id']:
            return jsonify({
                "success": False,
                "message": "Bu dosyaya erişim yetkiniz yok"
            }), 403
        
        # STEP dosya yolunu belirle
        step_path = None
        
        if analysis['file_type'] in ['step', 'stp']:
            # Direkt STEP dosyası
            step_path = analysis['file_path']
        elif analysis.get('matched_step_path'):
            # Eşleşmiş STEP dosyası
            step_path = analysis['matched_step_path']
        elif analysis.get('extracted_step_path'):
            # PDF'den çıkarılan STEP dosyası
            step_path = analysis['extracted_step_path']
        
        if not step_path or not os.path.exists(step_path):
            return jsonify({
                "success": False,
                "message": "STEP dosyası bulunamadı"
            }), 404
        
        print(f"[STL-GEN] 🔧 STL oluşturuluyor: {analysis_id}")
        
        # Session output directory
        session_output_dir = os.path.join("static", "stepviews", analysis_id)
        os.makedirs(session_output_dir, exist_ok=True)
        
        # STL dosya yolu
        stl_filename = f"model_{analysis_id}.stl"
        stl_path_full = os.path.join(session_output_dir, stl_filename)
        
        try:
            # CadQuery ile STEP'i import et ve STL olarak export et
            import cadquery as cq
            from cadquery import exporters
            
            # STEP dosyasını yükle
            assembly = cq.importers.importStep(step_path)
            shape = assembly.val()
            
            # STL olarak export et
            exporters.export(shape, stl_path_full)
            
            # Dosya boyutunu kontrol et
            if os.path.exists(stl_path_full):
                file_size = os.path.getsize(stl_path_full)
                print(f"[STL-GEN] ✅ STL oluşturuldu: {stl_filename} ({file_size} bytes)")
                
                # Analiz kaydını güncelle
                stl_relative = f"/static/stepviews/{analysis_id}/{stl_filename}"
                
                update_data = {
                    "stl_generated": True,
                    "stl_path": stl_relative,
                    "stl_file_size": file_size
                }
                
                # Enhanced renders'a ekle
                enhanced_renders = analysis.get('enhanced_renders', {})
                enhanced_renders['stl_model'] = {
                    "success": True,
                    "file_path": stl_relative,
                    "file_size": file_size,
                    "format": "stl"
                }
                update_data["enhanced_renders"] = enhanced_renders
                
                FileAnalysis.update_analysis(analysis_id, update_data)
                
                return jsonify({
                    "success": True,
                    "message": "STL dosyası başarıyla oluşturuldu",
                    "stl_path": stl_relative,
                    "stl_url": stl_relative,
                    "file_size": file_size,
                    "viewer_url": f"/step-viewer/{analysis_id}"
                }), 200
            else:
                raise Exception("STL dosyası oluşturulamadı")
                
        except Exception as stl_error:
            print(f"[STL-GEN] ❌ STL oluşturma hatası: {stl_error}")
            return jsonify({
                "success": False,
                "message": f"STL oluşturma hatası: {str(stl_error)}"
            }), 500
        
    except Exception as e:
        print(f"[STL-GEN] ❌ Beklenmeyen hata: {str(e)}")
        return jsonify({
            "success": False,
            "message": f"Beklenmeyen hata: {str(e)}"
        }), 500

# ===== STATUS AND MANAGEMENT ENDPOINTS =====

@upload_bp.route('/status/<analysis_id>', methods=['GET'])
@jwt_required()
def get_analysis_status(analysis_id):
    """Analiz durumunu getir"""
    try:
        current_user = get_current_user()
        
        analysis = FileAnalysis.find_by_id(analysis_id)
        if not analysis:
            return jsonify({
                "success": False,
                "message": "Analiz kaydı bulunamadı"
            }), 404
        
        # Kullanıcı yetkisi kontrolü
        if analysis['user_id'] != current_user['id']:
            return jsonify({
                "success": False,
                "message": "Bu dosyaya erişim yetkiniz yok"
            }), 403
        
        return jsonify({
            "success": True,
            "analysis": {
                "id": analysis['id'],
                "status": analysis.get('analysis_status', 'unknown'),
                "filename": analysis.get('original_filename'),
                "file_type": analysis.get('file_type'),
                "processing_time": analysis.get('processing_time'),
                "error_message": analysis.get('error_message'),
                "created_at": analysis.get('created_at'),
                "updated_at": analysis.get('updated_at'),
                "has_step_analysis": bool(analysis.get('step_analysis')),
                "has_renders": bool(analysis.get('enhanced_renders')),
                "material_matches_count": len(analysis.get('material_matches', [])),
                "render_count": len(analysis.get('enhanced_renders', {})),
                # Enhanced fields
                "has_matched_step": bool(analysis.get('matched_step_path')),
                "match_score": analysis.get('match_score'),
                "match_quality": analysis.get('match_quality'),
                "analysis_strategy": analysis.get('analysis_strategy'),
                "used_matched_step": analysis.get('used_matched_step', False),
                "step_source": analysis.get('step_source', 'none')
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Durum sorgulama hatası: {str(e)}"
        }), 500

@upload_bp.route('/my-uploads', methods=['GET'])
@jwt_required()
def get_my_uploads():
    """Kullanıcının yüklemelerini getir"""
    try:
        current_user = get_current_user()
        
        # Query parametreleri
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 20, type=int)
        file_type = request.args.get('file_type', '', type=str)
        status = request.args.get('status', '', type=str)
        
        skip = (page - 1) * limit
        
        # Kullanıcının analizlerini getir
        analyses = FileAnalysis.get_user_analyses(current_user['id'], limit, skip)
        total_count = FileAnalysis.get_user_analysis_count(current_user['id'])
        
        # Filtrele
        if file_type:
            analyses = [a for a in analyses if a.get('file_type') == file_type]
        if status:
            analyses = [a for a in analyses if a.get('analysis_status') == status]
        
        # Özet bilgiler ekle
        for analysis in analyses:
            analysis['summary'] = {
                "has_step_analysis": bool(analysis.get('step_analysis')),
                "has_renders": bool(analysis.get('enhanced_renders')),
                "material_count": len(analysis.get('material_matches', [])),
                "render_count": len(analysis.get('enhanced_renders', {})),
                "processing_time_formatted": f"{analysis.get('processing_time', 0):.2f}s" if analysis.get('processing_time') else "N/A",
                # Enhanced summary
                "has_matched_step": bool(analysis.get('matched_step_path')),
                "match_quality": analysis.get('match_quality', 'None'),
                "analysis_strategy": analysis.get('analysis_strategy', 'default')
            }
        
        return jsonify({
            "success": True,
            "uploads": analyses,
            "pagination": {
                "current_page": page,
                "total_pages": (total_count + limit - 1) // limit,
                "total_items": total_count,
                "items_per_page": limit
            },
            "filters_applied": {
                "file_type": file_type or None,
                "status": status or None
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Yüklemeler getirilemedi: {str(e)}"
        }), 500

@upload_bp.route('/delete/<analysis_id>', methods=['DELETE'])
@jwt_required()
def delete_analysis(analysis_id):
    """Analizi sil"""
    try:
        current_user = get_current_user()
        
        analysis = FileAnalysis.find_by_id(analysis_id)
        if not analysis:
            return jsonify({
                "success": False,
                "message": "Analiz kaydı bulunamadı"
            }), 404
        
        # Kullanıcı yetkisi kontrolü
        if analysis['user_id'] != current_user['id']:
            return jsonify({
                "success": False,
                "message": "Bu dosyayı silme yetkiniz yok"
            }), 403
        
        # Dosyaları sil
        try:
            if analysis.get('file_path') and os.path.exists(analysis['file_path']):
                os.remove(analysis['file_path'])
            
            # Render dosyalarını sil
            enhanced_renders = analysis.get('enhanced_renders', {})
            for view_name, view_data in enhanced_renders.items():
                if view_data.get('file_path'):
                    file_path = os.path.join(os.getcwd(), view_data['file_path'])
                    if os.path.exists(file_path):
                        os.remove(file_path)
        except Exception as file_error:
            print(f"[WARN] Dosya silme hatası: {file_error}")
        
        # Veritabanından sil
        success = FileAnalysis.delete_analysis(analysis_id)
        
        if success:
            return jsonify({
                "success": True,
                "message": "Analiz başarıyla silindi"
            }), 200
        else:
            return jsonify({
                "success": False,
                "message": "Analiz silinirken hata oluştu"
            }), 500
            
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Silme hatası: {str(e)}"
        }), 500

# ===== MATCHING ENDPOINTS =====

@upload_bp.route('/match-info/<analysis_id>', methods=['GET'])
@jwt_required()
def get_match_info(analysis_id):
    """PDF-STEP eşleştirme bilgilerini getir"""
    try:
        current_user = get_current_user()
        
        analysis = FileAnalysis.find_by_id(analysis_id)
        if not analysis:
            return jsonify({"success": False, "message": "Analiz bulunamadı"}), 404
        
        if analysis['user_id'] != current_user['id']:
            return jsonify({"success": False, "message": "Erişim yetkisi yok"}), 403
        
        match_info = {
            "analysis_id": analysis_id,
            "has_matched_step": bool(analysis.get('matched_step_file')),
            "match_details": None
        }
        
        if analysis.get('matched_step_file'):
            match_info["match_details"] = {
                "step_filename": analysis.get('matched_step_file'),
                "step_path": analysis.get('matched_step_path'),
                "match_score": analysis.get('match_score', 0),
                "match_quality": analysis.get('match_quality', 'Unknown'),
                "analysis_strategy": analysis.get('analysis_strategy', 'default'),
                "used_in_analysis": analysis.get('used_matched_step', False),
                "step_source": analysis.get('step_source', 'none')
            }
        
        return jsonify({
            "success": True,
            "match_info": match_info
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Match info error: {str(e)}"
        }), 500

@upload_bp.route('/re-match/<analysis_id>', methods=['POST'])
@jwt_required()
def re_match_analysis(analysis_id):
    """Analizi yeniden eşleştir"""
    try:
        current_user = get_current_user()
        
        analysis = FileAnalysis.find_by_id(analysis_id)
        if not analysis:
            return jsonify({"success": False, "message": "Analiz bulunamadı"}), 404
        
        if analysis['user_id'] != current_user['id']:
            return jsonify({"success": False, "message": "Erişim yetkisi yok"}), 403
        
        if analysis['file_type'] != 'pdf':
            return jsonify({"success": False, "message": "Sadece PDF dosyaları yeniden eşleştirilebilir"}), 400
        
        # Request body'den yeni STEP dosyası al
        data = request.get_json()
        new_step_analysis_id = data.get('step_analysis_id')
        
        if not new_step_analysis_id:
            return jsonify({"success": False, "message": "STEP analiz ID'si gerekli"}), 400
        
        # Yeni STEP analizini bul
        step_analysis = FileAnalysis.find_by_id(new_step_analysis_id)
        if not step_analysis:
            return jsonify({"success": False, "message": "STEP analizi bulunamadı"}), 404
        
        if step_analysis['user_id'] != current_user['id']:
            return jsonify({"success": False, "message": "STEP dosyasına erişim yetkisi yok"}), 403
        
        if step_analysis['file_type'] not in ['step', 'stp']:
            return jsonify({"success": False, "message": "Geçerli STEP dosyası değil"}), 400
        
        # Eşleştirme skorunu hesapla
        match_score = calculate_filename_similarity(
            analysis['original_filename'],
            step_analysis['original_filename']
        ) * 100
        
        # Analizi güncelle
        update_data = {
            "matched_step_file": step_analysis['filename'],
            "matched_step_path": step_analysis['file_path'],
            "match_score": round(match_score, 1),
            "match_quality": get_match_quality(match_score / 100),
            "analysis_strategy": "pdf_with_matched_step",
            "analysis_status": "uploaded"  # Yeniden analiz için
        }
        
        FileAnalysis.update_analysis(analysis_id, update_data)
        
        return jsonify({
            "success": True,
            "message": "Yeniden eşleştirme başarılı",
            "match_details": {
                "step_filename": step_analysis['original_filename'],
                "match_score": round(match_score, 1),
                "match_quality": get_match_quality(match_score / 100)
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Re-match error: {str(e)}"
        }), 500

# ===== BATCH ANALYSIS =====

@upload_bp.route('/batch-analyze', methods=['POST'])
@jwt_required()
def batch_analyze_enhanced():
    """✅ ENHANCED - Toplu analiz ve PDF-STEP eşleştirme desteği"""
    try:
        current_user = get_current_user()
        
        data = request.get_json()
        if not data or 'analysis_ids' not in data:
            return jsonify({
                "success": False,
                "message": "Analiz ID'leri gerekli"
            }), 400
        
        analysis_ids = data['analysis_ids']
        if not isinstance(analysis_ids, list) or len(analysis_ids) == 0:
            return jsonify({
                "success": False,
                "message": "Geçerli analiz ID listesi gerekli"
            }), 400
        
        if len(analysis_ids) > 20:
            return jsonify({
                "success": False,
                "message": "Maksimum 20 dosya aynı anda analiz edilebilir"
            }), 400
        
        print(f"[BATCH-ENHANCED] 📦 {len(analysis_ids)} dosya için toplu analiz başlatılıyor")
        
        # ✅ BATCH ANALYSIS LOGIC
        results = []
        
        for analysis_id in analysis_ids:
            try:
                analysis = FileAnalysis.find_by_id(analysis_id)
                if not analysis or analysis['user_id'] != current_user['id']:
                    results.append({
                        "analysis_id": analysis_id,
                        "status": "not_found_or_unauthorized",
                        "filename": None
                    })
                    continue
                
                # Analiz durumunu kontrol et
                current_status = analysis['analysis_status']
                if current_status in ['uploaded', 'failed']:
                    # Analizi başlat (ayrı thread'de veya queue'da)
                    # Şimdilik "queued" olarak işaretle
                    FileAnalysis.update_analysis(analysis_id, {
                        "analysis_status": "queued",
                        "batch_queued_at": time.time()
                    })
                    
                    results.append({
                        "analysis_id": analysis_id,
                        "status": "queued",
                        "filename": analysis.get('original_filename'),
                        "file_type": analysis.get('file_type'),
                        "has_matched_step": bool(analysis.get('matched_step_path')),
                        "analysis_strategy": analysis.get('analysis_strategy', 'default')
                    })
                elif current_status == 'analyzing':
                    results.append({
                        "analysis_id": analysis_id,
                        "status": "already_analyzing",
                        "filename": analysis.get('original_filename')
                    })
                else:
                    results.append({
                        "analysis_id": analysis_id,
                        "status": "already_processed",
                        "filename": analysis.get('original_filename')
                    })
                    
            except Exception as e:
                results.append({
                    "analysis_id": analysis_id,
                    "status": "error",
                    "error": str(e),
                    "filename": None
                })
        
        # Başarıyla kuyruğa alınan analizleri say
        queued_count = len([r for r in results if r['status'] == 'queued'])
        
        # ✅ OPTIONAL: Gerçek batch processing için background task başlat
        if queued_count > 0:
            batch_task_id = bg_processor.add_task(
                process_batch_analyses,
                args=(analysis_ids, current_user['id']),
                kwargs={}
            )
            
            print(f"[BATCH-ENHANCED] 🔄 Background batch processing started: {batch_task_id}")
        
        return jsonify({
            "success": True,
            "message": f"{queued_count} dosya için toplu analiz başlatıldı",
            "results": results,
            "summary": {
                "total_requested": len(analysis_ids),
                "queued": queued_count,
                "already_processed": len([r for r in results if r['status'] == 'already_processed']),
                "errors": len([r for r in results if r['status'] == 'error']),
                "pdf_with_step_count": len([r for r in results if r.get('has_matched_step', False)])
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Batch analiz hatası: {str(e)}"
        }), 500

# ===== EXPORT ENDPOINTS =====

@upload_bp.route('/export-excel/<analysis_id>', methods=['GET'])
@jwt_required()
def export_analysis_excel(analysis_id):
    """Analiz sonuçlarını Excel'e aktar (resimlerle birlikte)"""
    try:
        current_user = get_current_user()
        
        analysis = FileAnalysis.find_by_id(analysis_id)
        if not analysis:
            return jsonify({
                "success": False,
                "message": "Analiz kaydı bulunamadı"
            }), 404
        
        # Kullanıcı yetkisi kontrolü
        if analysis['user_id'] != current_user['id']:
            return jsonify({
                "success": False,
                "message": "Bu dosyaya erişim yetkiniz yok"
            }), 403
        
        try:
            import pandas as pd
            import io
            from datetime import datetime
            import os
            
            # STEP analizi verilerini topla
            step_analysis = analysis.get('step_analysis', {})
            
            # Malzeme bilgisini belirle
            material_matches = analysis.get('material_matches', [])
            material_name = "Bilinmiyor"
            
            if material_matches:
                first_match = material_matches[0]
                if isinstance(first_match, str) and "(" in first_match:
                    material_name = first_match.split("(")[0].strip()
                elif isinstance(first_match, str):
                    material_name = first_match
            
            # Excel satırı oluştur
            data = {
                "Ürün Görseli": "",  # Resim için boş bırak
                "Ürün Kodu": analysis.get('product_code', 'N/A'),
                "Dosya Adı": analysis.get('original_filename', 'N/A'),
                "Dosya Türü": analysis.get('file_type', 'N/A'),
                "Hammadde": material_name,
                "X+Pad (mm)": step_analysis.get('X+Pad (mm)', 0),
                "Y+Pad (mm)": step_analysis.get('Y+Pad (mm)', 0),
                "Z+Pad (mm)": step_analysis.get('Z+Pad (mm)', 0),
                "Silindirik Çap (mm)": step_analysis.get('Silindirik Çap (mm)', 0),
                "Ürün Hacmi (mm³)": step_analysis.get('Ürün Hacmi (mm³)', 0),
                "Toplam Yüzey Alanı (mm²)": step_analysis.get('Toplam Yüzey Alanı (mm²)', 0),
                "Hammadde Maliyeti (USD)": analysis.get('material_cost', 0),
                "Kütle (kg)": analysis.get('calculated_mass', 0),
                "Analiz Durumu": analysis.get('analysis_status', 'N/A'),
                "İşleme Süresi (s)": analysis.get('processing_time', 0),
                "Oluşturma Tarihi": analysis.get('created_at', 'N/A'),
                # Enhanced fields
                "Eşleşme Skoru": analysis.get('match_score', 'N/A'),
                "Eşleşme Kalitesi": analysis.get('match_quality', 'N/A'),
                "Analiz Stratejisi": analysis.get('analysis_strategy', 'N/A')
            }
            
            # Malzeme detayını ekle (varsa)
            if analysis.get('malzeme_detay'):
                data["Malzeme Eşleşmeleri"] = analysis['malzeme_detay']
            
            # Resim yolunu bul
            image_path = None
            enhanced_renders = analysis.get('enhanced_renders', {})
            
            # İzometrik görünüm varsa kullan
            if 'isometric' in enhanced_renders and enhanced_renders['isometric'].get('file_path'):
                image_path = enhanced_renders['isometric']['file_path']
            elif analysis.get('isometric_view_clean'):
                image_path = analysis['isometric_view_clean']
            elif analysis.get('isometric_view'):
                image_path = analysis['isometric_view']
            
            # Görsel yolunu tam path'e çevir
            if image_path:
                if image_path.startswith('/'):
                    image_path = image_path[1:]
                if not image_path.startswith('static'):
                    image_path = os.path.join('static', image_path)
                
                full_image_path = os.path.join(os.getcwd(), image_path)
                
                if not os.path.exists(full_image_path):
                    print(f"[EXPORT] ⚠️ Görsel dosyası bulunamadı: {full_image_path}")
                    image_path = None
                else:
                    print(f"[EXPORT] ✅ Görsel bulundu: {full_image_path}")
            
            # DataFrame oluştur
            df = pd.DataFrame([data])
            
            # Excel çıktısı (xlsxwriter ile)
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                # Ana sayfayı yaz
                df.to_excel(writer, sheet_name='Analiz Sonuçları', index=False, header=False, startrow=1)
                
                workbook = writer.book
                worksheet = writer.sheets['Analiz Sonuçları']
                
                # Sütun genişliklerini ayarla
                worksheet.set_column("A:A", 30)  # Görsel sütunu geniş
                worksheet.set_column("B:B", 20)  # Ürün Kodu
                worksheet.set_column("C:C", 25)  # Dosya Adı
                worksheet.set_column("D:D", 15)  # Dosya Türü
                worksheet.set_column("E:E", 20)  # Hammadde
                worksheet.set_column("F:Z", 18)  # Diğer sütunlar
                
                # Header stili
                header_format = workbook.add_format({
                    "bold": True,
                    "text_wrap": True,
                    "valign": "top",
                    "fg_color": "#D7E4BC",
                    "border": 1
                })
                
                # Header'ları yaz
                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num, value, header_format)
                
                # Resmi ekle
                if image_path and os.path.exists(full_image_path):
                    # Satır yüksekliğini artır
                    worksheet.set_row(1, 120)
                    
                    try:
                        # Resmi ekle
                        worksheet.insert_image("A2", full_image_path, {
                            "x_scale": 0.4,
                            "y_scale": 0.4,
                            "x_offset": 45,
                            "y_offset": 35
                        })
                        print(f"[EXPORT] ✅ Resim Excel'e eklendi: {image_path}")
                    except Exception as img_error:
                        print(f"[EXPORT] ❌ Resim ekleme hatası: {img_error}")
                
                # Ek sayfalar
                # Malzeme seçenekleri sayfası
                material_options = analysis.get('material_options', [])
                if material_options:
                    material_df = pd.DataFrame(material_options)
                    material_df.to_excel(writer, sheet_name='Malzeme Seçenekleri', index=False)
                
                # Enhanced renders sayfası
                if enhanced_renders:
                    renders_data = []
                    for view_name, view_data in enhanced_renders.items():
                        if view_data.get('success'):
                            renders_data.append({
                                "Görünüm": view_name,
                                "Dosya Yolu": view_data.get('file_path', ''),
                                "Başarılı": view_data.get('success', False),
                                "Format": view_data.get('format', 'png')
                            })
                    
                    if renders_data:
                        renders_df = pd.DataFrame(renders_data)
                        renders_df.to_excel(writer, sheet_name='3D Görünümler', index=False)
            
            output.seek(0)
            
            # Dosya adı oluştur
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"analiz_{analysis_id}_{timestamp}.xlsx"
            
            print(f"[EXPORT] ✅ Excel dosyası hazır: {filename}")
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except ImportError:
            return jsonify({
                "success": False,
                "message": "Excel export için pandas ve xlsxwriter gerekli"
            }), 500
        except Exception as excel_error:
            print(f"[EXPORT] ❌ Excel oluşturma hatası: {excel_error}")
            import traceback
            print(f"[EXPORT] 📋 Traceback: {traceback.format_exc()}")
            
            return jsonify({
                "success": False,
                "message": f"Excel oluşturma hatası: {str(excel_error)}"
            }), 500
            
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Excel export hatası: {str(e)}"
        }), 500

@upload_bp.route('/export-excel-multiple', methods=['POST'])
@jwt_required()
def export_multiple_analyses_excel():
    """Birden fazla analizi Excel'e aktar - KÜTLE VE MALİYET HESAPLAMALARİ İLE"""
    try:
        current_user = get_current_user()
        
        # Request body'den analysis_ids array'ini al
        data = request.get_json()
        if not data or 'analysis_ids' not in data:
            return jsonify({
                "success": False,
                "message": "analysis_ids array gerekli"
            }), 400
        
        analysis_ids = data['analysis_ids']
        if not isinstance(analysis_ids, list) or len(analysis_ids) == 0:
            return jsonify({
                "success": False,
                "message": "Geçerli analysis_ids array gerekli"
            }), 400
        
        if len(analysis_ids) > 50:  # Güvenlik limiti
            return jsonify({
                "success": False,
                "message": "Maksimum 50 analiz aynı anda export edilebilir"
            }), 400
        
        print(f"[EXCEL-MULTI] 📊 Çoklu Excel export başlıyor: {len(analysis_ids)} analiz")
        
        # Analizleri yükle ve yetki kontrolü
        analyses = []
        not_found = []
        unauthorized = []
        
        for analysis_id in analysis_ids:
            analysis = FileAnalysis.find_by_id(analysis_id)
            if not analysis:
                not_found.append(analysis_id)
                continue
            
            if analysis['user_id'] != current_user['id']:
                unauthorized.append(analysis_id)
                continue
            
            analyses.append(analysis)
        
        # Hata kontrolü
        if not_found:
            return jsonify({
                "success": False,
                "message": f"Bulunamayan analizler: {', '.join(not_found)}"
            }), 404
        
        if unauthorized:
            return jsonify({
                "success": False,
                "message": f"Yetkisiz erişim: {', '.join(unauthorized)}"
            }), 403
        
        if not analyses:
            return jsonify({
                "success": False,
                "message": "Export edilecek geçerli analiz bulunamadı"
            }), 400
        
        try:
            import pandas as pd
            import io
            from datetime import datetime
            import os
            
            print(f"[EXCEL-MULTI] ✅ {len(analyses)} analiz işlenecek")
            
            # Tüm analizler için enhanced veri hazırla
            excel_data = []
            total_calculated_mass = 0
            total_calculated_cost = 0
            successful_calculations = 0
            
            for analysis in analyses:
                print(f"[EXCEL-MULTI] 🔄 İşleniyor: {analysis.get('original_filename', 'unknown')}")
                
                # Her analiz için kütle ve maliyet hesapla
                calculated_data = calculate_mass_and_cost_for_analysis(analysis)
                
                # STEP analizi verilerini topla
                step_analysis = analysis.get('step_analysis', {})
                
                # Malzeme bilgisini belirle
                material_name = calculated_data['material_used']
                if material_name == 'Unknown':
                    material_matches = analysis.get('material_matches', [])
                    if material_matches:
                        first_match = material_matches[0]
                        if isinstance(first_match, str) and "(" in first_match:
                            material_name = first_match.split("(")[0].strip()
                        else:
                            material_name = str(first_match)
                
                # İşçilik ve toplam maliyet hesaplama
                calculated_mass_kg = calculated_data['calculated_mass_kg']
                calculated_material_cost = calculated_data['calculated_material_cost_usd']
                
                # İşçilik tahmini (kütle bazlı)
                estimated_labor_cost = 0
                if calculated_mass_kg > 0:
                    # Kütle bazlı işçilik: 0.5 kg altı = $10, üstü = kütle * $12
                    if calculated_mass_kg <= 0.5:
                        estimated_labor_cost = 10.0
                    else:
                        estimated_labor_cost = min(calculated_mass_kg * 12, 100.0)  # Max $100
                
                # Toplam birim maliyet
                unit_total_cost = calculated_material_cost + estimated_labor_cost
                
                # İstatistik için topla
                if calculated_mass_kg > 0:
                    total_calculated_mass += calculated_mass_kg
                    total_calculated_cost += unit_total_cost
                    successful_calculations += 1
                
                # Excel satırı oluştur - ENHANCED
                row_data = {
                    "Ürün Görseli": "",  # Resim için boş bırak - sonra eklenecek
                    "Analiz ID": analysis.get('id', 'N/A'),
                    "Dosya Adı": analysis.get('original_filename', 'N/A'),
                    "Dosya Türü": analysis.get('file_type', 'N/A'),
                    "Analiz Durumu": analysis.get('analysis_status', 'N/A'),
                    
                    # Malzeme bilgileri - ENHANCED
                    "Hammadde": material_name,
                    "Yoğunluk (g/cm³)": calculated_data['density_used'],
                    "Malzeme Fiyatı (USD/kg)": calculated_data['price_per_kg_used'],
                    
                    # Boyutlar
                    "X+Pad (mm)": step_analysis.get('X+Pad (mm)', step_analysis.get('X (mm)', 0)),
                    "Y+Pad (mm)": step_analysis.get('Y+Pad (mm)', step_analysis.get('Y (mm)', 0)),
                    "Z+Pad (mm)": step_analysis.get('Z+Pad (mm)', step_analysis.get('Z (mm)', 0)),
                    "Silindirik Çap (mm)": step_analysis.get('Silindirik Çap (mm)', 0),
                    
                    # Hacim ve kütle - HESAPLANMIŞ
                    "Hacim (mm³)": calculated_data['volume_used_mm3'],
                    "Ürün Hacmi (mm³)": step_analysis.get('Ürün Hacmi (mm³)', 0),
                    "Toplam Yüzey Alanı (mm²)": step_analysis.get('Toplam Yüzey Alanı (mm²)', 0),
                    "Kütle (kg)": calculated_mass_kg,  # HESAPLANMIŞ KÜTLE
                    
                    # Maliyet bilgileri - HESAPLANMIŞ
                    "Hammadde Maliyeti (USD)": calculated_material_cost,  # HESAPLANMIŞ MALİYET
                    "Tahmini İşçilik (USD)": round(estimated_labor_cost, 2),
                    "Birim Toplam Maliyet (USD)": round(unit_total_cost, 2),
                    
                    # Enhanced matching fields
                    "Eşleşme Skoru": analysis.get('match_score', 'N/A'),
                    "Eşleşme Kalitesi": analysis.get('match_quality', 'N/A'),
                    "Analiz Stratejisi": analysis.get('analysis_strategy', 'N/A'),
                    "Eşleşmiş STEP Kullanıldı": "Evet" if analysis.get('used_matched_step', False) else "Hayır",
                    
                    # Meta veriler
                    "İşleme Süresi (s)": analysis.get('processing_time', 0),
                    "Oluşturma Tarihi": analysis.get('created_at', 'N/A'),
                    "Render Sayısı": len(analysis.get('enhanced_renders', {})),
                    "PDF'den STEP": "Evet" if analysis.get('pdf_step_extracted', False) else "Hayır"
                }
                
                # Malzeme detayını ekle (varsa)
                if analysis.get('material_matches'):
                    row_data["Malzeme Eşleşmeleri"] = "; ".join(analysis['material_matches'][:3])  # İlk 3'ü
                
                # Resim yolunu bul ve ekle
                image_path = None
                enhanced_renders = analysis.get('enhanced_renders', {})
                
                # İzometrik görünüm varsa kullan
                if 'isometric' in enhanced_renders and enhanced_renders['isometric'].get('file_path'):
                    image_path = enhanced_renders['isometric']['file_path']
                elif analysis.get('isometric_view_clean'):
                    image_path = analysis['isometric_view_clean']
                elif analysis.get('isometric_view'):
                    image_path = analysis['isometric_view']
                
                # Görsel yolunu tam path'e çevir
                full_image_path = None
                if image_path:
                    if image_path.startswith('/'):
                        image_path = image_path[1:]
                    if not image_path.startswith('static'):
                        image_path = os.path.join('static', image_path)
                    
                    full_image_path = os.path.join(os.getcwd(), image_path)
                    
                    if not os.path.exists(full_image_path):
                        print(f"[EXCEL-MULTI] ⚠️ Görsel dosyası bulunamadı: {full_image_path}")
                        full_image_path = None
                    else:
                        print(f"[EXCEL-MULTI] ✅ Görsel bulundu: {full_image_path}")
                
                # Row data'ya image path'i ekle (Excel'de kullanılacak)
                row_data["_image_path"] = full_image_path
                
                excel_data.append(row_data)
                
                print(f"[EXCEL-MULTI] ✅ {analysis.get('original_filename')}: {calculated_mass_kg:.3f} kg, ${calculated_material_cost:.2f}")
            
            # DataFrame oluştur
            df = pd.DataFrame(excel_data)
            
            # _image_path sütununu DataFrame'den çıkar (sadece internal kullanım için)
            image_paths = df["_image_path"].tolist()
            df = df.drop(columns=["_image_path"])
            
            print(f"[EXCEL-MULTI] 📋 DataFrame oluşturuldu: {len(df)} satır")
            print(f"[EXCEL-MULTI] 📊 Toplam kütle: {total_calculated_mass:.3f} kg")
            print(f"[EXCEL-MULTI] 💰 Toplam maliyet: ${total_calculated_cost:.2f}")
            
            # Excel çıktısı (xlsxwriter ile)
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                # Ana sayfayı yaz
                df.to_excel(writer, sheet_name='Analiz Sonuçları', index=False, header=False, startrow=1)
                
                workbook = writer.book
                worksheet = writer.sheets['Analiz Sonuçları']
                
                # Sütun genişliklerini ayarla
                column_widths = {
                    0: 60,   # Görsel sütunu geniş
                    1: 15,   # Analiz ID
                    2: 25,   # Dosya Adı
                    3: 12,   # Dosya Türü
                    4: 15,   # Analiz Durumu
                    5: 20,   # Hammadde
                    6: 12,   # Yoğunluk
                    7: 15,   # Malzeme Fiyatı
                    8: 12,   # X+Pad
                    9: 12,   # Y+Pad
                    10: 12,  # Z+Pad
                    11: 15,  # Silindirik Çap
                    12: 15,  # Hacim
                    13: 15,  # Ürün Hacmi
                    14: 18,  # Yüzey Alanı
                    15: 12,  # Kütle
                    16: 18,  # Hammadde Maliyeti
                    17: 15,  # İşçilik
                    18: 18,  # Birim Toplam
                    19: 15,  # Eşleşme Skoru
                    20: 15,  # Eşleşme Kalitesi
                    21: 18,  # Analiz Stratejisi
                    22: 15,  # Eşleşmiş STEP
                    23: 15,  # İşleme Süresi
                    24: 20,  # Tarih
                    25: 12,  # Render Sayısı
                    26: 12,  # PDF STEP
                    27: 25   # Malzeme Eşleşmeleri
                }
                
                for col_index, width in column_widths.items():
                    if col_index < len(df.columns):
                        col_letter = chr(65 + col_index) if col_index < 26 else chr(64 + col_index // 26) + chr(65 + col_index % 26)
                        worksheet.set_column(f"{col_letter}:{col_letter}", width)
                
                # Header stili
                header_format = workbook.add_format({
                    "bold": True,
                    "text_wrap": True,
                    "valign": "top",
                    "fg_color": "#D7E4BC",
                    "border": 1,
                    "font_size": 10
                })
                
                # Sayısal değer formatları
                number_format = workbook.add_format({'num_format': '#,##0.000'})
                currency_format = workbook.add_format({'num_format': '$#,##0.00'})
                
                # Header'ları yaz
                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num, value, header_format)
                
                # Resimleri satırlara ekle
                for row_idx, image_path in enumerate(image_paths):
                    excel_row = row_idx + 1  # +1 çünkü header var
                    
                    # Satır yüksekliğini artır (resim için)
                    worksheet.set_row(excel_row, 120)
                    
                    if image_path and os.path.exists(image_path):
                        try:
                            # Resmi ekle (optimized boyutlarda)
                            worksheet.insert_image(f"A{excel_row + 1}", image_path, {
                                "x_scale": 0.35,
                                "y_scale": 0.35,
                                "x_offset": 5,
                                "y_offset": 5
                            })
                            print(f"[EXCEL-MULTI] 🖼️ Satır {excel_row + 1}: Resim eklendi")
                        except Exception as img_error:
                            print(f"[EXCEL-MULTI] ❌ Satır {excel_row + 1} resim ekleme hatası: {img_error}")
                            worksheet.write(f"A{excel_row + 1}", "Resim Hatası")
                    else:
                        worksheet.write(f"A{excel_row + 1}", "Resim Yok")
                
                # Sayısal sütunlara format uygula
                # Kütle sütunu (kg)
                mass_col = None
                cost_cols = []
                
                for col_idx, col_name in enumerate(df.columns):
                    if "Kütle" in col_name:
                        mass_col = col_idx
                    elif any(keyword in col_name for keyword in ["Maliyet", "İşçilik", "Toplam", "Fiyat"]):
                        cost_cols.append(col_idx)
                
                # Kütle formatı
                if mass_col is not None:
                    col_letter = chr(65 + mass_col)
                    worksheet.set_column(f"{col_letter}:{col_letter}", 12, number_format)
                
                # Para formatı
                for col_idx in cost_cols:
                    col_letter = chr(65 + col_idx)
                    worksheet.set_column(f"{col_letter}:{col_letter}", 15, currency_format)
                
                # Ek sayfalar
                
                # 1. Malzeme özeti sayfası
                material_summary = {}
                for analysis in analyses:
                    calculated_data = calculate_mass_and_cost_for_analysis(analysis)
                    material = calculated_data['material_used']
                    
                    if material not in material_summary:
                        material_summary[material] = {
                            'count': 0,
                            'total_mass': 0,
                            'total_cost': 0,
                            'density': calculated_data['density_used'],
                            'price_per_kg': calculated_data['price_per_kg_used']
                        }
                    
                    material_summary[material]['count'] += 1
                    material_summary[material]['total_mass'] += calculated_data['calculated_mass_kg']
                    material_summary[material]['total_cost'] += calculated_data['calculated_material_cost_usd']
                
                if material_summary:
                    summary_data = []
                    for material, data in material_summary.items():
                        summary_data.append({
                            'Malzeme': material,
                            'Parça Sayısı': data['count'],
                            'Toplam Kütle (kg)': round(data['total_mass'], 3),
                            'Toplam Maliyet (USD)': round(data['total_cost'], 2),
                            'Ortalama Kütle (kg)': round(data['total_mass'] / data['count'], 3),
                            'Yoğunluk (g/cm³)': data['density'],
                            'Fiyat (USD/kg)': data['price_per_kg']
                        })
                    
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='Malzeme Özeti', index=False)
                    print(f"[EXCEL-MULTI] 📄 Malzeme özeti sayfası: {len(summary_data)} malzeme")
                
                # 2. Eşleştirme istatistikleri sayfası
                matching_stats = {
                    "Metrik": [
                        "Toplam PDF Dosyası",
                        "Eşleşmiş PDF-STEP Çifti",
                        "Eşleşme Başarı Oranı (%)",
                        "Excellent Eşleşme",
                        "Good Eşleşme", 
                        "Fair Eşleşme",
                        "Poor Eşleşme",
                        "Ortalama Eşleşme Skoru",
                        "Eşleşmiş STEP Kullanım Oranı (%)"
                    ]
                }
                
                pdf_analyses = [a for a in analyses if a.get('file_type') == 'pdf']
                matched_analyses = [a for a in pdf_analyses if a.get('matched_step_path')]
                
                match_qualities = {}
                total_match_score = 0
                match_score_count = 0
                used_matched_step_count = 0
                
                for analysis in matched_analyses:
                    quality = analysis.get('match_quality', 'Unknown')
                    match_qualities[quality] = match_qualities.get(quality, 0) + 1
                    
                    score = analysis.get('match_score', 0)
                    if score:
                        total_match_score += score
                        match_score_count += 1
                    
                    if analysis.get('used_matched_step', False):
                        used_matched_step_count += 1
                
                matching_stats["Değer"] = [
                    len(pdf_analyses),
                    len(matched_analyses),
                    round((len(matched_analyses) / len(pdf_analyses) * 100), 1) if pdf_analyses else 0,
                    match_qualities.get('Excellent', 0),
                    match_qualities.get('Good', 0),
                    match_qualities.get('Fair', 0),
                    match_qualities.get('Poor', 0),
                    round(total_match_score / match_score_count, 1) if match_score_count else 0,
                    round((used_matched_step_count / len(matched_analyses) * 100), 1) if matched_analyses else 0
                ]
                
                matching_df = pd.DataFrame(matching_stats)
                matching_df.to_excel(writer, sheet_name='Eşleştirme İstatistikleri', index=False)
                
                # 3. Genel istatistikler sayfası
                stats_data = {
                    "Metrik": [
                        "Toplam Analiz Sayısı",
                        "Başarılı Kütle Hesaplaması", 
                        "Başarısız Analizler",
                        "STEP Dosyaları",
                        "PDF Dosyaları",
                        "PDF'den STEP Çıkarılan",
                        "Ortalama İşleme Süresi (s)",
                        "Toplam Kütle (kg)",
                        "Toplam Hammadde Maliyeti (USD)",
                        "Ortalama Birim Maliyet (USD)"
                    ],
                    "Değer": [
                        len(analyses),
                        successful_calculations,
                        len([a for a in analyses if a.get('analysis_status') == 'failed']),
                        len([a for a in analyses if a.get('file_type') in ['step', 'stp']]),
                        len([a for a in analyses if a.get('file_type') == 'pdf']),
                        len([a for a in analyses if a.get('pdf_step_extracted', False)]),
                        round(sum([a.get('processing_time', 0) for a in analyses]) / len(analyses), 2),
                        round(total_calculated_mass, 3),
                        round(sum([calculate_mass_and_cost_for_analysis(a)['calculated_material_cost_usd'] for a in analyses]), 2),
                        round(total_calculated_cost / len(analyses), 2) if analyses else 0
                    ]
                }
                
                stats_df = pd.DataFrame(stats_data)
                stats_df.to_excel(writer, sheet_name='İstatistikler', index=False)
                print(f"[EXCEL-MULTI] 📊 İstatistik sayfası oluşturuldu")
                
                # 4. Detaylı malzeme hesaplamaları sayfası
                detailed_calcs = []
                for analysis in analyses:
                    calc_data = calculate_mass_and_cost_for_analysis(analysis)
                    detailed_calcs.append({
                        'Analiz ID': analysis.get('id'),
                        'Dosya Adı': analysis.get('original_filename'),
                        'Malzeme': calc_data['material_used'],
                        'Hacim (mm³)': calc_data['volume_used_mm3'],
                        'Yoğunluk (g/cm³)': calc_data['density_used'],
                        'Kütle (kg)': calc_data['calculated_mass_kg'],
                        'Fiyat (USD/kg)': calc_data['price_per_kg_used'],
                        'Maliyet (USD)': calc_data['calculated_material_cost_usd'],
                        'Hesaplama Formülü': f"{calc_data['volume_used_mm3']} mm³ × {calc_data['density_used']} g/cm³ ÷ 1,000,000 = {calc_data['calculated_mass_kg']} kg"
                    })
                
                if detailed_calcs:
                    detailed_df = pd.DataFrame(detailed_calcs)
                    detailed_df.to_excel(writer, sheet_name='Hesaplama Detayları', index=False)
                    print(f"[EXCEL-MULTI] 🧮 Hesaplama detayları sayfası: {len(detailed_calcs)} hesaplama")
            
            output.seek(0)
            
            # Dosya adı oluştur
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"coklu_analiz_{len(analyses)}_dosya_{timestamp}.xlsx"
            
            print(f"[EXCEL-MULTI] ✅ Excel dosyası hazır: {filename}")
            print(f"[EXCEL-MULTI] 📈 Başarılı hesaplamalar: {successful_calculations}/{len(analyses)}")
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except ImportError:
            return jsonify({
                "success": False,
                "message": "Excel export için pandas ve xlsxwriter gerekli"
            }), 500
        except Exception as excel_error:
            print(f"[EXCEL-MULTI] ❌ Excel oluşturma hatası: {excel_error}")
            import traceback
            print(f"[EXCEL-MULTI] 📋 Traceback: {traceback.format_exc()}")
            
            return jsonify({
                "success": False,
                "message": f"Excel oluşturma hatası: {str(excel_error)}"
            }), 500
            
    except Exception as e:
        print(f"[EXCEL-MULTI] ❌ Genel hata: {str(e)}")
        return jsonify({
            "success": False,
            "message": f"Çoklu Excel export hatası: {str(e)}"
        }), 500

# ===== MERGE WITH EXCEL =====

@upload_bp.route('/merge-with-excel', methods=['POST'])
@jwt_required()
def merge_with_excel():
    """Excel dosyasını analiz sonuçlarıyla birleştir - KÜTLE VE FİYAT HESAPLAMALARİ İLE"""
    try:
        current_user = get_current_user()
        
        # Form verilerini kontrol et
        if 'excel_file' not in request.files:
            return jsonify({
                "success": False,
                "message": "Excel dosyası bulunamadı"
            }), 400
        
        excel_file = request.files['excel_file']
        analysis_ids = request.form.getlist('analysis_ids')
        
        if excel_file.filename == '':
            return jsonify({
                "success": False,
                "message": "Excel dosyası seçilmedi"
            }), 400
        
        if not analysis_ids:
            return jsonify({
                "success": False,
                "message": "Analiz ID'leri belirtilmedi"
            }), 400
        
        print(f"[MERGE] 📊 Excel birleştirme başlıyor: {excel_file.filename}")
        print(f"[MERGE] 🔢 Analiz ID'leri: {analysis_ids}")
        
        # Excel dosyası kontrolü
        if not excel_file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({
                "success": False,
                "message": "Sadece Excel dosyaları (.xlsx, .xls) desteklenir"
            }), 400
        
        # Analizleri yükle ve yetki kontrolü
        analyses = []
        for analysis_id in analysis_ids:
            analysis = FileAnalysis.find_by_id(analysis_id)
            if not analysis:
                return jsonify({
                    "success": False,
                    "message": f"Analiz bulunamadı: {analysis_id}"
                }), 404
            
            if analysis['user_id'] != current_user['id']:
                return jsonify({
                    "success": False,
                    "message": f"Analiz erişim yetkisi yok: {analysis_id}"
                }), 403
            
            analyses.append(analysis)
        
        print(f"[MERGE] ✅ {len(analyses)} analiz yüklendi")
        
        # Excel işleme
        try:
            import openpyxl
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
            import re
            import math
            import io
            from datetime import datetime
            
            # Excel dosyasını yükle
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            print(f"[MERGE] ✅ Excel yüklendi. Satır: {ws.max_row}, Sütun: {ws.max_column}")
            
            # Geliştirilmiş normalize fonksiyonu
            def normalize_robust(text):
                if not text:
                    return ""
                
                if not isinstance(text, str):
                    text = str(text)
                
                # Türkçe karakterleri çevir
                replacements = {
                    'ç': 'c', 'ğ': 'g', 'ı': 'i', 'ö': 'o', 'ş': 's', 'ü': 'u',
                    'Ç': 'C', 'Ğ': 'G', 'İ': 'I', 'Ö': 'O', 'Ş': 'S', 'Ü': 'U'
                }
                for tr_char, en_char in replacements.items():
                    text = text.replace(tr_char, en_char)
                
                normalized = re.sub(r'[^\w]', '', text.lower())
                return normalized
            
            def extract_numbers(text):
                if not text:
                    return []
                numbers = re.findall(r'\d+', str(text))
                return numbers
            
            # Header analizi ve sütun tespiti
            header_row = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
            print(f"[MERGE] 📋 Header satırı: {header_row}")
            
            # Malzeme No sütununu bul
            malzeme_no_patterns = [
                "malzeme no", "malzemeno", "malzeme_no", "malzeme numarası", "malzeme numarasi",
                "ürün kodu", "urun kodu", "ürün no", "urun no", "kod", "no", "part", "item"
            ]
            
            malzeme_col_index = None
            for i, header in enumerate(header_row):
                if header:
                    normalized_header = normalize_robust(header)
                    print(f"[MERGE] 🔍 Header {i+1}: '{header}' -> '{normalized_header}'")
                    
                    for pattern in malzeme_no_patterns:
                        if normalize_robust(pattern) == normalized_header:
                            malzeme_col_index = i + 1  # 1-based
                            print(f"[MERGE] ✅ Malzeme No sütunu: '{header}' (sütun {malzeme_col_index})")
                            break
                    if malzeme_col_index:
                        break
            
            if not malzeme_col_index:
                malzeme_col_index = 3
                print(f"[MERGE] ⚠️ Malzeme No sütunu bulunamadı, sütun {malzeme_col_index} kullanılıyor")
            
            # İhale miktarı sütununu bul
            ihale_col_index = None
            ihale_patterns = ["ihale", "miktar", "adet", "quantity", "amount"]
            
            for i, header in enumerate(header_row):
                if header:
                    normalized_header = normalize_robust(header)
                    for pattern in ihale_patterns:
                        if pattern in normalized_header:
                            ihale_col_index = i + 1
                            print(f"[MERGE] ✅ İhale sütunu: '{header}' (sütun {ihale_col_index})")
                            break
                    if ihale_col_index:
                        break
            
            if not ihale_col_index:
                ihale_col_index = malzeme_col_index + 1
                print(f"[MERGE] ⚠️ İhale sütunu bulunamadı, sütun {ihale_col_index} kullanılıyor")
            
            # İhale sütunundan sonraki sütunları sil
            columns_to_keep = ihale_col_index
            columns_to_delete = ws.max_column - columns_to_keep
            
            for _ in range(columns_to_delete):
                if ws.max_column > columns_to_keep:
                    ws.delete_cols(columns_to_keep + 1)
            
            print(f"[MERGE] 🗑️ {columns_to_delete} sütun silindi")
            
            # Yeni sütun başlıkları ekle
            new_headers = [
                "Ürün Görseli", "Hammadde", "X+Pad (mm)", "Y+Pad (mm)", "Z+Pad (mm)",
                "Silindirik Çap (mm)", "Kütle (kg)", "Hammadde Maliyeti (USD)",
                "Kaplama", "Helicoil", "Markalama", "İşçilik", "Birim Fiyat", "Toplam",
                "Eşleşme Skoru", "Analiz Stratejisi"  # Enhanced columns
            ]
            
            start_col = columns_to_keep + 1
            for i, header in enumerate(new_headers):
                ws.cell(row=1, column=start_col + i, value=header)
            
            # Sütun genişlikleri
            for i in range(len(new_headers)):
                col_letter = openpyxl.utils.get_column_letter(start_col + i)
                if i == 0:  # Görsel sütunu
                    ws.column_dimensions[col_letter].width = 25
                else:
                    ws.column_dimensions[col_letter].width = 14
            
            # Analiz verilerini lookup tablosu hazırla - ENHANCED MATERIAL CALCULATIONS
            analysis_lookup = {}
            
            for analysis in analyses:
                # Product code çıkarma stratejileri
                product_codes = []
                
                # 1. Direkt product_code alanından
                if analysis.get('product_code'):
                    product_codes.append(str(analysis['product_code']))
                
                # 2. Filename'den rakam çıkarma
                filename = analysis.get('original_filename', '')
                if filename:
                    # Başından rakam çıkar
                    front_numbers = re.findall(r'^\d+', filename)
                    if front_numbers:
                        product_codes.append(front_numbers[0])
                    
                    # Tüm rakamları çıkar
                    all_numbers = re.findall(r'\d+', filename)
                    product_codes.extend(all_numbers)
                
                # 3. Analysis ID'yi de ekle
                product_codes.append(str(analysis.get('id', '')))
                
                # Kütle ve fiyat hesaplamaları
                analysis_calculated_data = calculate_mass_and_cost_for_analysis(analysis)
                
                # Benzersiz kodları normalize et ve ekle
                for code in set(product_codes):
                    if code and len(code) >= 3:  # En az 3 karakter
                        normalized_code = normalize_robust(code)
                        if normalized_code:
                            # Analysis'e hesaplanmış verileri ekle
                            enhanced_analysis = analysis.copy()
                            enhanced_analysis.update(analysis_calculated_data)
                            
                            analysis_lookup[normalized_code] = enhanced_analysis
                            print(f"[MERGE] 📝 Lookup eklendi: '{code}' -> '{normalized_code}' -> {analysis['id']} (kütle: {analysis_calculated_data.get('calculated_mass_kg', 'N/A')} kg)")
            
            print(f"[MERGE] 📋 Toplam lookup entries: {len(analysis_lookup)}")
            
            # Satırları işle ve eşleştir
            matched_count = 0
            total_rows = 0
            
            for row in range(2, ws.max_row + 1):
                total_rows += 1
                
                # Excel'den malzeme numarasını al
                malzeme_cell = ws.cell(row=row, column=malzeme_col_index).value
                
                if not malzeme_cell:
                    print(f"[MERGE] ⚠️ Satır {row}: Malzeme numarası boş")
                    continue
                
                excel_malzeme = str(malzeme_cell).strip()
                print(f"[MERGE] 🔍 Satır {row}: Excel malzeme = '{excel_malzeme}'")
                
                # Eşleşmeyi bul
                matched_analysis = None
                match_method = ""
                
                # 1. Tam eşleşme
                excel_normalized = normalize_robust(excel_malzeme)
                if excel_normalized in analysis_lookup:
                    matched_analysis = analysis_lookup[excel_normalized]
                    match_method = "exact"
                
                # 2. Kısmi eşleşme (başından)
                if not matched_analysis:
                    for lookup_code, analysis in analysis_lookup.items():
                        if excel_normalized.startswith(lookup_code) or lookup_code.startswith(excel_normalized):
                            if len(lookup_code) >= 4:  # Minimum güvenlik
                                matched_analysis = analysis
                                match_method = "partial_start"
                                break
                
                # 3. Sayısal eşleşme
                if not matched_analysis:
                    excel_numbers = extract_numbers(excel_malzeme)
                    for lookup_code, analysis in analysis_lookup.items():
                        lookup_numbers = extract_numbers(lookup_code)
                        if excel_numbers and lookup_numbers:
                            # En büyük sayıları karşılaştır
                            if max(excel_numbers) == max(lookup_numbers):
                                matched_analysis = analysis
                                match_method = "numeric"
                                break
                
                # Eşleşme bulunursa verileri yaz
                if matched_analysis:
                    matched_count += 1
                    print(f"[MERGE] ✅ Satır {row}: '{excel_malzeme}' eşleşti -> {matched_analysis['id']} ({match_method})")
                    
                    # Hesaplanmış verileri al
                    step_analysis = matched_analysis.get('step_analysis', {})
                    
                    # Malzeme bilgisi
                    material_matches = matched_analysis.get('material_matches', [])
                    material_name = matched_analysis.get('material_used', 'Bilinmiyor')
                    
                    if not material_name or material_name == 'Bilinmiyor':
                        if material_matches:
                            first_match = material_matches[0]
                            if isinstance(first_match, str) and "(" in first_match:
                                material_name = first_match.split("(")[0].strip()
                            else:
                                material_name = str(first_match)
                    
                    # Hesaplanmış kütle ve maliyet - LOOKUP'TAN AL
                    kutle_kg = matched_analysis.get('calculated_mass_kg', 0)
                    maliyet_usd = matched_analysis.get('calculated_material_cost_usd', 0)
                    density_used = matched_analysis.get('density_used', 2.7)
                    price_per_kg_used = matched_analysis.get('price_per_kg_used', 4.5)
                    
                    # İşçilik maliyeti hesaplama (basit tahmin)
                    iscilik_usd = 0
                    if kutle_kg > 0:
                        # Kütle bazlı işçilik tahmini: büyük parça = daha fazla işçilik
                        iscilik_base = min(kutle_kg * 15, 50)  # Max $50
                        iscilik_usd = round(iscilik_base, 2)
                    
                    # Birim fiyat hesaplama (hammadde + işçilik)
                    birim_fiyat = maliyet_usd + iscilik_usd
                    
                    # İhale miktarını al (Toplam hesaplama için)
                    ihale_miktari = 1  # Default
                    ihale_cell = ws.cell(row=row, column=ihale_col_index).value
                    if ihale_cell:
                        try:
                            # Virgülü noktaya çevir ve sayıya dönüştür
                            ihale_str = str(ihale_cell).replace(',', '.')
                            ihale_miktari = float(ihale_str)
                        except:
                            ihale_miktari = 1
                    
                    # Toplam hesaplama
                    toplam_maliyet = birim_fiyat * ihale_miktari
                    
                    values_data = [
                        None,  # Görsel (sonra eklenecek)
                        material_name,
                        step_analysis.get("X+Pad (mm)", 0) or step_analysis.get("X (mm)", 0),
                        step_analysis.get("Y+Pad (mm)", 0) or step_analysis.get("Y (mm)", 0),
                        step_analysis.get("Z+Pad (mm)", 0) or step_analysis.get("Z (mm)", 0),
                        step_analysis.get("Silindirik Çap (mm)", 0) or step_analysis.get("Çap (mm)", 0),
                        kutle_kg if kutle_kg > 0 else None,           # HESAPLANMIŞ KÜTLE
                        maliyet_usd if maliyet_usd > 0 else None,     # HESAPLANMIŞ MALİYET
                        "",  # Kaplama - boş bırak
                        "",  # Helicoil - boş bırak
                        "",  # Markalama - boş bırak
                        iscilik_usd if iscilik_usd > 0 else "",      # İşçilik
                        birim_fiyat if birim_fiyat > 0 else "",      # Birim Fiyat
                        toplam_maliyet if toplam_maliyet > 0 else "", # Toplam
                        matched_analysis.get('match_score', 'N/A'),   # Eşleşme Skoru
                        matched_analysis.get('analysis_strategy', 'N/A')  # Analiz Stratejisi
                    ]
                    
                    print(f"[MERGE] 📊 Satır {row} değerler:")
                    print(f"   - Kütle: {kutle_kg} kg (density: {density_used} g/cm³)")
                    print(f"   - Hammadde Maliyeti: ${maliyet_usd} (${price_per_kg_used}/kg)")
                    print(f"   - İşçilik: ${iscilik_usd}")
                    print(f"   - Birim Fiyat: ${birim_fiyat}")
                    print(f"   - İhale Miktarı: {ihale_miktari}")
                    print(f"   - Toplam: ${toplam_maliyet}")
                    
                    # Satır yüksekliğini ayarla
                    ws.row_dimensions[row].height = 120
                    
                    # Verileri hücrelere yaz
                    for i, value in enumerate(values_data):
                        target_col = start_col + i
                        target_cell = ws.cell(row=row, column=target_col)
                        
                        if i == 0:  # Görsel sütunu
                            # Görseli bul ve ekle
                            image_path = None
                            enhanced_renders = matched_analysis.get('enhanced_renders', {})
                            
                            # Görsel kaynak önceliği
                            if 'isometric' in enhanced_renders and enhanced_renders['isometric'].get('file_path'):
                                image_path = enhanced_renders['isometric']['file_path']
                            elif matched_analysis.get('isometric_view_clean'):
                                image_path = matched_analysis['isometric_view_clean']
                            elif matched_analysis.get('isometric_view'):
                                image_path = matched_analysis['isometric_view']
                            
                            if image_path:
                                # Path'i düzelt
                                if image_path.startswith('/'):
                                    image_path = image_path[1:]
                                if not image_path.startswith('static'):
                                    image_path = os.path.join('static', image_path)
                                
                                full_image_path = os.path.join(os.getcwd(), image_path)
                                
                                if os.path.exists(full_image_path):
                                    try:
                                        img = XLImage(full_image_path)
                                        
                                        # Güvenli boyutlandırma
                                        max_width = 160
                                        max_height = 100
                                        
                                        if img.width > 0 and img.height > 0:
                                            # Aspect ratio koru
                                            width_ratio = max_width / img.width
                                            height_ratio = max_height / img.height
                                            scale_ratio = min(width_ratio, height_ratio)
                                            
                                            img.width = int(img.width * scale_ratio)
                                            img.height = int(img.height * scale_ratio)
                                        
                                        # Hücre koordinatını hesapla
                                        cell_coord = f"{openpyxl.utils.get_column_letter(target_col)}{row}"
                                        ws.add_image(img, cell_coord)
                                        
                                        print(f"[MERGE] 🖼️ Satır {row}: Resim eklendi ({img.width}x{img.height})")
                                        
                                    except Exception as img_error:
                                        print(f"[MERGE] ❌ Satır {row} resim hatası: {img_error}")
                                        target_cell.value = "Resim Hatası"
                                else:
                                    print(f"[MERGE] ⚠️ Satır {row}: Resim dosyası bulunamadı: {full_image_path}")
                                    target_cell.value = "Resim Bulunamadı"
                            else:
                                target_cell.value = "Resim Yok"
                        else:
                            # Sayısal değerleri formatla ve yaz
                            if isinstance(value, (float, int)) and value is not None:
                                if value != 0:  # Sıfır değerleri yazma
                                    if isinstance(value, float):
                                        # Para birimi sütunları için 2 decimal
                                        if i in [7, 11, 12, 13]:  # Maliyet, İşçilik, Birim Fiyat, Toplam
                                            target_cell.value = round(value, 2)
                                            target_cell.number_format = '#,##0.00'
                                        # Kütle için 3 decimal
                                        elif i == 6:  # Kütle
                                            target_cell.value = round(value, 3)
                                            target_cell.number_format = '#,##0.000'
                                        # Boyutlar için 1 decimal
                                        elif i in [2, 3, 4, 5]:  # Boyutlar
                                            target_cell.value = round(value, 1)
                                            target_cell.number_format = '#,##0.0'
                                        else:
                                            target_cell.value = round(value, 2)
                                    else:
                                        target_cell.value = value
                                        if i in [7, 11, 12, 13]:  # Para sütunları
                                            target_cell.number_format = '#,##0.00'
                            elif value and str(value).strip():  # Boş olmayan string değerler
                                target_cell.value = str(value).strip()
                        
                        # Hücre hizalaması
                        target_cell.alignment = Alignment(
                            horizontal='center',
                            vertical='center',
                            wrap_text=True
                        )
                
                else:
                    print(f"[MERGE] ❌ Satır {row}: '{excel_malzeme}' eşleşmedi")
            
            print(f"[MERGE] 📊 İşlem tamamlandı: {matched_count}/{total_rows} eşleşme")
            
            # Header stillendirme
            header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
            header_font = Font(bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=1, column=col)
                header_cell.fill = header_fill
                header_cell.font = header_font
                header_cell.border = border
                header_cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
            
            # Dosyayı kaydet ve döndür
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Dosya adı oluştur
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            original_name = excel_file.filename.rsplit('.', 1)[0]
            filename = f"{original_name}_merged_{timestamp}.xlsx"
            
            print(f"[MERGE] ✅ Excel başarıyla birleştirildi: {filename}")
            print(f"[MERGE] 📈 Sonuç: {matched_count}/{total_rows} satır eşleşti")
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except ImportError as e:
            missing_lib = str(e).split("'")[1] if "'" in str(e) else str(e)
            return jsonify({
                "success": False,
                "message": f"Gerekli kütüphane bulunamadı: {missing_lib}. pip install {missing_lib} çalıştırın."
            }), 500
        except Exception as excel_error:
            print(f"[MERGE] ❌ Excel işleme hatası: {excel_error}")
            import traceback
            print(f"[MERGE] 📋 Traceback: {traceback.format_exc()}")
            
            return jsonify({
                "success": False,
                "message": f"Excel işleme hatası: {str(excel_error)}",
                "details": traceback.format_exc()
            }), 500
    
    except Exception as e:
        print(f"[MERGE] ❌ Genel hata: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            "success": False,
            "message": f"Birleştirme hatası: {str(e)}"
        }), 500

# ===== UTILITY AND STATISTICS ENDPOINTS =====

@upload_bp.route('/search', methods=['GET'])
@jwt_required()
def search_analyses():
    """Analizlerde arama yap"""
    try:
        current_user = get_current_user()
        
        # Query parametreleri
        search_term = request.args.get('q', '', type=str)
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 20, type=int)
        
        if not search_term or len(search_term.strip()) < 2:
            return jsonify({
                "success": False,
                "message": "Arama terimi en az 2 karakter olmalı"
            }), 400
        
        # Arama yap
        results = FileAnalysis.search_analyses(current_user['id'], search_term.strip())
        
        # Pagination
        total = len(results)
        start = (page - 1) * limit
        end = start + limit
        paginated_results = results[start:end]
        
        # Özet bilgiler ekle
        for result in paginated_results:
            result['search_relevance'] = {
                "filename_match": search_term.lower() in result.get('original_filename', '').lower(),
                "material_match": any(search_term.lower() in m.lower() for m in result.get('material_matches', [])),
                "file_type_match": search_term.lower() == result.get('file_type', '').lower(),
                "has_matched_step": bool(result.get('matched_step_path'))  # Enhanced field
            }
        
        return jsonify({
            "success": True,
            "search_results": paginated_results,
            "search_term": search_term,
            "pagination": {
                "current_page": page,
                "total_pages": (total + limit - 1) // limit,
                "total_items": total,
                "items_per_page": limit
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Arama hatası: {str(e)}"
        }), 500

@upload_bp.route('/statistics', methods=['GET'])
@jwt_required()
def get_user_statistics():
    """Kullanıcının dosya istatistikleri"""
    try:
        current_user = get_current_user()
        
        # Kullanıcının tüm analizlerini al
        all_analyses = FileAnalysis.get_user_analyses(current_user['id'], limit=1000)
        
        # İstatistikleri hesapla
        stats = {
            "total_files": len(all_analyses),
            "by_status": {},
            "by_file_type": {},
            "total_processing_time": 0,
            "successful_analyses": 0,
            "failed_analyses": 0,
            "files_with_renders": 0,
            "total_materials_found": 0,
            # Enhanced matching statistics
            "pdf_files": 0,
            "step_files": 0,
            "matched_pdf_step_pairs": 0,
            "unmatched_pdfs": 0,
            "excellent_matches": 0,
            "good_matches": 0,
            "fair_matches": 0,
            "poor_matches": 0,
            "used_matched_step_count": 0,
            "average_match_score": 0
        }
        
        total_match_score = 0
        match_count = 0
        
        for analysis in all_analyses:
            # Durum istatistikleri
            status = analysis.get('analysis_status', 'unknown')
            stats['by_status'][status] = stats['by_status'].get(status, 0) + 1
            
            # Dosya türü istatistikleri
            file_type = analysis.get('file_type', 'unknown')
            stats['by_file_type'][file_type] = stats['by_file_type'].get(file_type, 0) + 1
            
            if file_type == 'pdf':
                stats['pdf_files'] += 1
            elif file_type in ['step', 'stp']:
                stats['step_files'] += 1
            
            # İşleme süresi
            processing_time = analysis.get('processing_time', 0)
            if processing_time:
                stats['total_processing_time'] += processing_time
            
            # Başarı oranları
            if status == 'completed':
                stats['successful_analyses'] += 1
            elif status == 'failed':
                stats['failed_analyses'] += 1
            
            # Render sayısı
            if analysis.get('enhanced_renders'):
                stats['files_with_renders'] += 1
            
            # Malzeme sayısı
            materials = analysis.get('material_matches', [])
            stats['total_materials_found'] += len(materials)
            
            # Enhanced matching statistics
            if analysis.get('matched_step_path'):
                stats['matched_pdf_step_pairs'] += 1
                
                match_quality = analysis.get('match_quality', '').lower()
                if match_quality == 'excellent':
                    stats['excellent_matches'] += 1
                elif match_quality == 'good':
                    stats['good_matches'] += 1
                elif match_quality == 'fair':
                    stats['fair_matches'] += 1
                elif match_quality == 'poor':
                    stats['poor_matches'] += 1
                
                match_score = analysis.get('match_score', 0)
                if match_score:
                    total_match_score += match_score
                    match_count += 1
                
                if analysis.get('used_matched_step', False):
                    stats['used_matched_step_count'] += 1
            elif file_type == 'pdf':
                stats['unmatched_pdfs'] += 1
        
        # Ortalamalar
        stats['average_processing_time'] = stats['total_processing_time'] / max(1, stats['successful_analyses'])
        stats['success_rate'] = (stats['successful_analyses'] / max(1, stats['total_files'])) * 100
        stats['average_match_score'] = round(total_match_score / max(1, match_count), 1)
        stats['matching_success_rate'] = (stats['matched_pdf_step_pairs'] / max(1, stats['pdf_files'])) * 100
        stats['matched_step_usage_rate'] = (stats['used_matched_step_count'] / max(1, stats['matched_pdf_step_pairs'])) * 100
        
        return jsonify({
            "success": True,
            "statistics": stats,
            "user_id": current_user['id']
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"İstatistik hatası: {str(e)}"
        }), 500

@upload_bp.route('/matching-stats', methods=['GET'])
@jwt_required()
def get_matching_statistics():
    """PDF-STEP eşleştirme istatistikleri"""
    try:
        current_user = get_current_user()
        
        # Kullanıcının tüm analizlerini al
        all_analyses = FileAnalysis.get_user_analyses(current_user['id'], limit=1000)
        
        # Eşleştirme istatistikleri
        stats = {
            "total_analyses": len(all_analyses),
            "pdf_files": len([a for a in all_analyses if a.get('file_type') == 'pdf']),
            "step_files": len([a for a in all_analyses if a.get('file_type') in ['step', 'stp']]),
            "matched_pairs": len([a for a in all_analyses if a.get('matched_step_file')]),
            "match_quality_distribution": {
                "excellent": 0,
                "good": 0,
                "fair": 0,
                "poor": 0
            },
            "average_match_score": 0,
            "successful_step_usage": len([a for a in all_analyses if a.get('used_matched_step', False)]),
            "analysis_strategy_distribution": {},
            "step_source_distribution": {}
        }
        
        # Match quality hesaplama
        matched_analyses = [a for a in all_analyses if a.get('match_quality')]
        total_score = 0
        
        for analysis in matched_analyses:
            quality = analysis.get('match_quality', '').lower()
            if quality in stats['match_quality_distribution']:
                stats['match_quality_distribution'][quality] += 1
            
            score = analysis.get('match_score', 0)
            if score:
                total_score += score
            
            # Analysis strategy distribution
            strategy = analysis.get('analysis_strategy', 'unknown')
            stats['analysis_strategy_distribution'][strategy] = stats['analysis_strategy_distribution'].get(strategy, 0) + 1
            
            # Step source distribution
            step_source = analysis.get('step_source', 'none')
            stats['step_source_distribution'][step_source] = stats['step_source_distribution'].get(step_source, 0) + 1
        
        if matched_analyses:
            stats['average_match_score'] = round(total_score / len(matched_analyses), 1)
        
        # Success rates
        stats['matching_success_rate'] = (stats['matched_pairs'] / max(1, stats['pdf_files'])) * 100
        stats['step_usage_rate'] = (stats['successful_step_usage'] / max(1, stats['matched_pairs'])) * 100
        
        return jsonify({
            "success": True,
            "matching_statistics": stats
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Matching stats error: {str(e)}"
        }), 500

@upload_bp.route('/batch-status', methods=['GET'])
@jwt_required()
def get_batch_status():
    """Batch işlem durumunu getir"""
    try:
        current_user = get_current_user()
        
        # Kullanıcının kuyruktaki ve işlemdeki analizlerini bul
        all_analyses = FileAnalysis.get_user_analyses(current_user['id'], limit=100)
        
        batch_status = {
            "queued": [a for a in all_analyses if a.get('analysis_status') == 'queued'],
            "analyzing": [a for a in all_analyses if a.get('analysis_status') == 'analyzing'],
            "completed_recently": [
                a for a in all_analyses 
                if a.get('analysis_status') == 'completed' 
                and a.get('batch_processed', False)
                and (time.time() - a.get('updated_at', 0)) < 3600  # Son 1 saat
            ]
        }
        
        return jsonify({
            "success": True,
            "batch_status": {
                "queued_count": len(batch_status['queued']),
                "analyzing_count": len(batch_status['analyzing']),
                "completed_recently_count": len(batch_status['completed_recently']),
                "details": batch_status
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Batch status error: {str(e)}"
        }), 500

@upload_bp.route('/supported-formats', methods=['GET'])
def get_supported_formats():
    """✅ ENHANCED - Desteklenen dosya formatları (CAD desteği ile)"""
    return jsonify({
        "success": True,
        "supported_formats": {
            "upload": list(ALLOWED_EXTENSIONS),
            "analysis": {
                "pdf": "PDF doküman analizi ve malzeme tanıma",
                "doc": "Word doküman analizi",
                "docx": "Word doküman analizi", 
                "step": "3D STEP dosya analizi ve rendering",
                "stp": "3D STEP dosya analizi ve rendering",
                "prt": "NX Part dosya analizi (STEP'e çevrilerek)",  # ✅ YENİ
                "catpart": "CATIA Part dosya analizi (STEP'e çevrilerek)"  # ✅ YENİ
            }
        },
        "limits": {
            "max_file_size_mb": MAX_FILE_SIZE // (1024 * 1024),
            "max_files_per_request": MAX_FILES_PER_REQUEST
        },
        "features": {
            "material_recognition": True,
            "3d_rendering": True,
            "cost_estimation": True,
            "wireframe_generation": True,
            "technical_drawings": True,
            "excel_export": True,
            "pdf_cad_matching": True,  # ✅ ENHANCED
            "batch_processing": True,
            "cad_conversion": {  # ✅ YENİ
                "prt_to_step": True,
                "catpart_to_step": True,
                "auto_conversion": True,
                "conversion_engine": "FreeCAD"
            }
        },
        "cad_conversion_info": {  # ✅ YENİ
            "supported_input_formats": ["prt", "catpart"],
            "output_format": "step",
            "conversion_engine": "FreeCAD",
            "freecad_available": cad_converter.freecad_path is not None,
            "freecad_path": cad_converter.freecad_path,
            "estimated_conversion_time": "30-120 seconds per file"
        }
    }), 200

@upload_bp.route('/download/<analysis_id>/<view_type>', methods=['GET'])
@jwt_required()
def download_render(analysis_id, view_type):
    """Render dosyasını indir"""
    try:
        current_user = get_current_user()
        
        analysis = FileAnalysis.find_by_id(analysis_id)
        if not analysis:
            return jsonify({
                "success": False,
                "message": "Analiz kaydı bulunamadı"
            }), 404
        
        # Kullanıcı yetkisi kontrolü
        if analysis['user_id'] != current_user['id']:
            return jsonify({
                "success": False,
                "message": "Bu dosyaya erişim yetkiniz yok"
            }), 403
        
        # Render dosyasını bul
        enhanced_renders = analysis.get('enhanced_renders', {})
        if view_type not in enhanced_renders:
            return jsonify({
                "success": False,
                "message": f"'{view_type}' görünümü bulunamadı"
            }), 404
        
        render_data = enhanced_renders[view_type]
        if not render_data.get('success') or not render_data.get('file_path'):
            return jsonify({
                "success": False,
                "message": f"'{view_type}' dosyası mevcut değil"
            }), 404
        
        file_path = os.path.join(os.getcwd(), render_data['file_path'])
        if not os.path.exists(file_path):
            return jsonify({
                "success": False,
                "message": "Dosya sistemde bulunamadı"
            }), 404
        
        # Dosyayı indir
        filename = f"{analysis.get('original_filename', 'render')}_{view_type}.png"
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='image/png'
        )
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"İndirme hatası: {str(e)}"
        }), 500

@upload_bp.route('/performance-stats', methods=['GET'])
@jwt_required()
def get_performance_stats():
    """Get performance statistics"""
    try:
        # Background task queue stats
        queue_size = bg_processor.task_queue.qsize()
        completed_tasks = len(bg_processor.results)
        
        # System stats (basic)
        try:
            import psutil
            cpu_percent = psutil.cpu_percent(interval=1)
            memory_percent = psutil.virtual_memory().percent
        except ImportError:
            cpu_percent = 0
            memory_percent = 0
        
        return jsonify({
            "success": True,
            "performance": {
                "background_queue_size": queue_size,
                "completed_background_tasks": completed_tasks,
                "cpu_usage_percent": cpu_percent,
                "memory_usage_percent": memory_percent,
                "optimization_status": "active",
                "pdf_step_matching_enabled": True,
                "enhanced_analysis_enabled": True
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Performance stats error: {str(e)}"
        }), 500

# ===== ENHANCED BACKGROUND RENDERING =====

def background_render_task_enhanced(analysis_id: str, step_path: str, analysis_strategy: str = "default"):
    """Enhanced background rendering with strategy awareness - FIXED VERSION"""
    try:
        print(f"[BG-RENDER-ENHANCED] 🎨 Enhanced background render başlıyor: {analysis_id}")
        print(f"[BG-RENDER-ENHANCED] 📂 STEP path: {step_path}")
        print(f"[BG-RENDER-ENHANCED] 📋 Strategy: {analysis_strategy}")
        
        from services.step_renderer import StepRendererEnhanced
        from models.file_analysis import FileAnalysis
        
        # ✅ DOSYA VARLIK KONTROLÜ
        if not os.path.exists(step_path):
            error_msg = f"STEP dosyası bulunamadı: {step_path}"
            print(f"[BG-RENDER-ENHANCED] ❌ {error_msg}")
            
            FileAnalysis.update_analysis(analysis_id, {
                "render_status": "failed",
                "render_error": error_msg
            })
            return {"success": False, "error": error_msg}
        
        step_renderer = StepRendererEnhanced()
        
        # Strategy-based render quality
        render_quality = "high" if analysis_strategy == "pdf_with_matched_step" else "medium"
        include_materials = analysis_strategy != "step_only"
        
        print(f"[BG-RENDER-ENHANCED] ⚙️ Render ayarları:")
        print(f"   - Quality: {render_quality}")
        print(f"   - Include Materials: {include_materials}")
        
        # ✅ RENDER GENERATION
        render_result = step_renderer.generate_comprehensive_views(
            step_path,
            analysis_id=analysis_id,
            include_dimensions=True,
            include_materials=include_materials,
            high_quality=(render_quality == "high")
        )
        
        print(f"[BG-RENDER-ENHANCED] 📊 Render result: {render_result.get('success', False)}")
        
        if render_result.get('success', False):
            renders = render_result.get('renders', {})
            print(f"[BG-RENDER-ENHANCED] 🖼️ Generated renders: {list(renders.keys())}")
            
            # ✅ DATABASE UPDATE WITH DETAILED LOGGING
            update_data = {
                "enhanced_renders": renders,  # Bu key field!
                "render_quality": render_quality,
                "render_status": "completed",
                "render_strategy": analysis_strategy,
                "render_count": len(renders),
                "last_render_update": time.time()
            }
            
            # Add main isometric view
            if 'isometric' in renders:
                isometric_data = renders['isometric']
                if isometric_data.get('success'):
                    update_data["isometric_view"] = isometric_data.get('file_path')
                    if isometric_data.get('excel_path'):
                        update_data["isometric_view_clean"] = isometric_data.get('excel_path')
                    print(f"[BG-RENDER-ENHANCED] 🎯 Isometric view: {isometric_data.get('file_path')}")
            
            # ✅ STL GENERATION (conditional based on strategy)
            if analysis_strategy in ["pdf_with_matched_step", "step_only"]:
                try:
                    print(f"[BG-RENDER-ENHANCED] 🔧 STL generation başlıyor...")
                    
                    import cadquery as cq
                    from cadquery import exporters
                    
                    session_output_dir = os.path.join("static", "stepviews", analysis_id)
                    os.makedirs(session_output_dir, exist_ok=True)
                    
                    stl_filename = f"model_{analysis_id}.stl"
                    stl_path_full = os.path.join(session_output_dir, stl_filename)
                    
                    # STL generation with error handling
                    assembly = cq.importers.importStep(step_path)
                    shape = assembly.val()
                    exporters.export(shape, stl_path_full)
                    
                    if os.path.exists(stl_path_full):
                        file_size = os.path.getsize(stl_path_full)
                        stl_relative = f"/static/stepviews/{analysis_id}/{stl_filename}"
                        
                        update_data.update({
                            "stl_generated": True,
                            "stl_path": stl_relative,
                            "stl_file_size": file_size
                        })
                        
                        # STL'i enhanced_renders'a da ekle
                        update_data["enhanced_renders"]["stl_model"] = {
                            "success": True,
                            "file_path": stl_relative,
                            "file_size": file_size,
                            "format": "stl"
                        }
                        
                        print(f"[BG-RENDER-ENHANCED] ✅ STL generated: {stl_filename} ({file_size} bytes)")
                        
                except Exception as stl_error:
                    print(f"[BG-RENDER-ENHANCED] ⚠️ STL generation failed: {stl_error}")
                    update_data["stl_generation_error"] = str(stl_error)
            
            # ✅ CRITICAL: DATABASE UPDATE WITH ERROR HANDLING
            print(f"[BG-RENDER-ENHANCED] 💾 Database update başlıyor...")
            print(f"[BG-RENDER-ENHANCED] 📊 Update data keys: {list(update_data.keys())}")
            print(f"[BG-RENDER-ENHANCED] 🖼️ Enhanced renders count: {len(update_data.get('enhanced_renders', {}))}")
            
            try:
                success = FileAnalysis.update_analysis(analysis_id, update_data)
                if success:
                    print(f"[BG-RENDER-ENHANCED] ✅ Database update successful")
                    
                    # ✅ VERIFICATION: Re-read and verify
                    verification = FileAnalysis.find_by_id(analysis_id)
                    if verification:
                        verified_renders = verification.get('enhanced_renders', {})
                        print(f"[BG-RENDER-ENHANCED] ✅ Verification: {len(verified_renders)} renders in DB")
                        if len(verified_renders) != len(renders):
                            print(f"[BG-RENDER-ENHANCED] ⚠️ Render count mismatch: Expected {len(renders)}, Got {len(verified_renders)}")
                    else:
                        print(f"[BG-RENDER-ENHANCED] ❌ Verification failed: Analysis not found")
                else:
                    print(f"[BG-RENDER-ENHANCED] ❌ Database update failed")
                    return {"success": False, "error": "Database update failed"}
                    
            except Exception as db_error:
                print(f"[BG-RENDER-ENHANCED] ❌ Database update exception: {db_error}")
                import traceback
                print(f"[BG-RENDER-ENHANCED] 📋 DB Traceback: {traceback.format_exc()}")
                return {"success": False, "error": f"Database error: {str(db_error)}"}
            
            print(f"[BG-RENDER-ENHANCED] ✅ Enhanced render completed: {analysis_id} - {len(renders)} views")
            return {
                "success": True, 
                "renders": len(renders), 
                "strategy": analysis_strategy,
                "render_paths": list(renders.keys())
            }
            
        else:
            # Render failed
            error_msg = render_result.get('message', 'Enhanced render hatası')
            print(f"[BG-RENDER-ENHANCED] ❌ Render failed: {error_msg}")
            
            FileAnalysis.update_analysis(analysis_id, {
                "render_status": "failed",
                "render_error": error_msg
            })
            
            return {"success": False, "error": error_msg}
            
    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback_str = traceback.format_exc()
        
        print(f"[BG-RENDER-ENHANCED] ❌ Enhanced background render error: {error_msg}")
        print(f"[BG-RENDER-ENHANCED] 📋 Traceback: {traceback_str}")
        
        try:
            FileAnalysis.update_analysis(analysis_id, {
                "render_status": "failed",
                "render_error": error_msg,
                "render_traceback": traceback_str
            })
        except Exception as db_error:
            print(f"[BG-RENDER-ENHANCED] ❌ Failed to update error status: {db_error}")
            
        return {"success": False, "error": error_msg}

# ===== RENDER STATUS CHECK ENDPOINT - ENHANCED =====


def process_batch_analyses(analysis_ids: List[str], user_id: str):
    """Background batch processing function"""
    try:
        print(f"[BATCH-PROCESS] 🔄 Processing {len(analysis_ids)} analyses for user {user_id}")
        
        processed = 0
        errors = 0
        
        for analysis_id in analysis_ids:
            try:
                analysis = FileAnalysis.find_by_id(analysis_id)
                if analysis and analysis['user_id'] == user_id and analysis['analysis_status'] == 'queued':
                    print(f"[BATCH-PROCESS] 📄 Processing: {analysis['original_filename']}")
                    
                    # Update status to analyzing
                    FileAnalysis.update_analysis(analysis_id, {"analysis_status": "analyzing"})
                    
                    # TODO: In real implementation, call the actual analysis function
                    # For now, simulate processing
                    time.sleep(2)
                    
                    # Mark as completed
                    FileAnalysis.update_analysis(analysis_id, {
                        "analysis_status": "completed",
                        "processing_time": 2.0,
                        "batch_processed": True
                    })
                    
                    processed += 1
                    
            except Exception as e:
                print(f"[BATCH-PROCESS] ❌ Error processing {analysis_id}: {e}")
                FileAnalysis.update_analysis(analysis_id, {
                    "analysis_status": "failed",
                    "error_message": f"Batch processing error: {str(e)}"
                })
                errors += 1
        
        print(f"[BATCH-PROCESS] ✅ Batch completed: {processed} processed, {errors} errors")
        return {"processed": processed, "errors": errors}
        
    except Exception as e:
        print(f"[BATCH-PROCESS] ❌ Batch processing failed: {e}")
        return {"error": str(e)}

def calculate_mass_and_cost_for_analysis(analysis):
    """Analiz için kütle ve maliyet hesaplama - ENHANCED"""
    try:
        # Default değerler
        result = {
            'calculated_mass_kg': 0.0,
            'calculated_material_cost_usd': 0.0,
            'density_used': 2.7,
            'price_per_kg_used': 4.5,
            'volume_used_mm3': 0.0,
            'material_used': 'Unknown'
        }
        
        # STEP analizinden hacim al
        step_analysis = analysis.get('step_analysis', {})
        volume_mm3 = 0
        
        # Hacim kaynaklarını dene
        if step_analysis.get('Prizma Hacmi (mm³)'):
            volume_mm3 = step_analysis['Prizma Hacmi (mm³)']
        elif step_analysis.get('Ürün Hacmi (mm³)'):
            volume_mm3 = step_analysis['Ürün Hacmi (mm³)']
        elif step_analysis.get('volume_mm3'):
            volume_mm3 = step_analysis['volume_mm3']
        
        if volume_mm3 <= 0:
            print(f"[CALC-MASS] ⚠️ Analiz {analysis.get('id', 'unknown')}: Geçerli hacim bulunamadı")
            return result
        
        result['volume_used_mm3'] = volume_mm3
        
        # Malzeme bilgisini belirle - ENHANCED
        material_matches = analysis.get('material_matches', [])
        material_name = 'Unknown'
        best_confidence = 0
        
        # En yüksek confidence'a sahip malzemeyi bul
        if material_matches:
            best_material = None
            
            for match in material_matches:
                if isinstance(match, str):
                    # Confidence değerini çıkar
                    confidence_match = re.search(r'%(\d+)', match)
                    if confidence_match:
                        confidence_value = int(confidence_match.group(1))
                    elif "estimated" in match.lower():
                        confidence_value = 70
                    else:
                        confidence_value = 50
                    
                    # En yüksek confidence'ı bul
                    if confidence_value > best_confidence:
                        best_confidence = confidence_value
                        best_material = match
            
            # En iyi malzemeyi seç
            if best_material:
                if "(" in best_material:
                    material_name = best_material.split("(")[0].strip()
                else:
                    material_name = best_material.strip()
                
                print(f"[CALC-MASS] 🏆 En iyi malzeme seçildi: {material_name} (%{best_confidence})")
            else:
                # Fallback: ilk malzemeyi kullan
                first_match = material_matches[0]
                if isinstance(first_match, str) and "(" in first_match:
                    material_name = first_match.split("(")[0].strip()
                else:
                    material_name = str(first_match) if first_match else 'Unknown'
        
        result['material_used'] = material_name
        
        # MongoDB'den malzeme verilerini al
        try:
            from utils.database import db
            database = db.get_db()
            
            # Malzeme ara
            material = database.materials.find_one({
                "$or": [
                    {"name": {"$regex": f"^{material_name}$", "$options": "i"}},
                    {"name": {"$regex": material_name, "$options": "i"}},
                    {"aliases": {"$in": [material_name]}},
                    {"aliases": {"$elemMatch": {"$regex": material_name, "$options": "i"}}}
                ]
            })
            
            if material:
                density = material.get("density", 2.7)
                price_per_kg = material.get("price_per_kg", 4.5)
                print(f"[CALC-MASS] ✅ MongoDB'de bulundu: {material.get('name')} (density: {density}, price: ${price_per_kg})")
            else:
                print(f"[CALC-MASS] ⚠️ MongoDB'de bulunamadı: {material_name}, varsayılan kullanılıyor")
                # Varsayılan değerler - yaygın malzemeler için
                if "6061" in material_name.upper():
                    density, price_per_kg = 2.7, 4.5
                elif "7075" in material_name.upper():
                    density, price_per_kg = 2.81, 6.2
                elif "304" in material_name.upper():
                    density, price_per_kg = 7.93, 8.5
                elif "316" in material_name.upper():
                    density, price_per_kg = 7.98, 12.0
                elif "ST37" in material_name.upper() or "S235" in material_name.upper():
                    density, price_per_kg = 7.85, 2.2
                else:
                    density, price_per_kg = 2.7, 4.5
            
        except Exception as db_error:
            print(f"[CALC-MASS] ❌ MongoDB hatası: {db_error}")
            density, price_per_kg = 2.7, 4.5
        
        result['density_used'] = density
        result['price_per_kg_used'] = price_per_kg
        
        # Kütle hesaplama
        mass_kg = (volume_mm3 * density) / 1_000_000
        result['calculated_mass_kg'] = round(mass_kg, 3)
        
        # Maliyet hesaplama
        material_cost_usd = mass_kg * price_per_kg
        result['calculated_material_cost_usd'] = round(material_cost_usd, 2)
        
        print(f"[CALC-MASS] ✅ Hesaplama tamamlandı: {volume_mm3} mm³ x {density} g/cm³ = {mass_kg:.3f} kg x ${price_per_kg} = ${material_cost_usd:.2f}")
        print(f"[CALC-MASS] 🎯 Seçilen malzeme: {material_name} (confidence: %{best_confidence})")
        
        return result
        
    except Exception as e:
        import traceback
        print(f"[CALC-MASS] ❌ Kütle/maliyet hesaplama hatası: {e}")
        print(f"[CALC-MASS] 📋 Traceback: {traceback.format_exc()}")
        return {
            'calculated_mass_kg': 0.0,
            'calculated_material_cost_usd': 0.0,
            'density_used': 2.7,
            'price_per_kg_used': 4.5,
            'volume_used_mm3': 0.0,
            'material_used': 'Unknown'
        }
                
@upload_bp.route('/cad-conversion-status', methods=['GET'])
@jwt_required()
def get_cad_conversion_status():
    """✅ YENİ - CAD conversion sisteminin durumunu kontrol et"""
    try:
        return jsonify({
            "success": True,
            "cad_conversion": {
                "available": cad_converter.freecad_path is not None,
                "freecad_path": cad_converter.freecad_path,
                "supported_formats": list(cad_converter.supported_formats.keys()),
                "temp_directory": cad_converter.temp_dir,
                "temp_files_count": len(list(Path(cad_converter.temp_dir).glob("*"))) if os.path.exists(cad_converter.temp_dir) else 0
            },
            "system_requirements": {
                "freecad_required": True,
                "freecad_min_version": "0.19",
                "python_modules": ["subprocess", "tempfile", "pathlib"]
            },
            "conversion_statistics": {
                "temp_directory_size_mb": sum(f.stat().st_size for f in Path(cad_converter.temp_dir).glob("**/*") if f.is_file()) / (1024*1024) if os.path.exists(cad_converter.temp_dir) else 0
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"CAD conversion status error: {str(e)}"
        }), 500

# ===== CAD CONVERSION CLEANUP ENDPOINT =====

@upload_bp.route('/cad-cleanup', methods=['POST'])
@jwt_required()
def cleanup_cad_temp_files():
    """✅ YENİ - CAD conversion geçici dosyalarını temizle"""
    try:
        current_user = get_current_user()
        
        # Admin yetkisi kontrol et (opsiyonel)
        # if current_user.get('role') != 'admin':
        #     return jsonify({"success": False, "message": "Admin yetkisi gerekli"}), 403
        
        max_age_hours = request.json.get('max_age_hours', 24) if request.json else 24
        
        removed_count = cad_converter.cleanup_temp_files(max_age_hours)
        
        return jsonify({
            "success": True,
            "message": f"{removed_count} geçici dosya temizlendi",
            "removed_files": removed_count,
            "max_age_hours": max_age_hours
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Cleanup error: {str(e)}"
        }), 500
